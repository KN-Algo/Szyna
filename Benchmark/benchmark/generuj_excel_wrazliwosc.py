# generuj_excel_wrazliwosc.py
#
# Buduje Excel z wyników test_wrazliwosc_dwie_lokalizacje.py
# (WRAZLIWOSC_ZBIORCZY.csv -> Podsumowanie_wrazliwosc.xlsx):
#   - "Wyniki"            - surowa tabela, jeden wiersz na (lokalizacja, algorytm, scenariusz, szum)
#   - "Odchylenie_energii" - % zmiany energii względem scenariusza 'nominal' (bez szumu) TEGO SAMEGO
#                             algorytmu/lokalizacji - pokazuje kto jest odporny na zaburzenie transmitancji
#   - "Jakosc_autotestu"  - błąd identyfikacji K/T1/L (%) dla algorytmów adaptacyjnych, osobno z/bez szumu
#
# Uruchomienie:
#   python generuj_excel_wrazliwosc.py
# albo programowo: generuj_excel_wrazliwosc.main(sciezka_csv, sciezka_xlsx)

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_WYNIKOW = os.environ.get(
    'SZYNA_FOLDER_WYNIKOW_WRAZ', os.path.join(BASE_DIR, "wyniki", "wrazliwosc_2lokalizacje"))
SCIEZKA_CSV_DOMYSLNA = os.path.join(FOLDER_WYNIKOW, "WRAZLIWOSC_ZBIORCZY.csv")
SCIEZKA_XLSX_DOMYSLNA = os.path.join(FOLDER_WYNIKOW, "Podsumowanie_wrazliwosc.xlsx")

FONT_NAZWA = 'Arial'
KOLOR_NAGLOWEK_BG = '1F4E78'
KOLOR_NAGLOWEK_FG = 'FFFFFF'
FONT_NAGLOWEK = Font(name=FONT_NAZWA, bold=True, color=KOLOR_NAGLOWEK_FG, size=11)
FILL_NAGLOWEK = PatternFill('solid', fgColor=KOLOR_NAGLOWEK_BG)
FONT_ZWYKLY = Font(name=FONT_NAZWA, size=10)
FONT_POGRUBIONY = Font(name=FONT_NAZWA, bold=True, size=10)
WYROWNANIE_SRODEK = Alignment(horizontal='center', vertical='center')
OBRAMOWANIE_CIENKIE = Border(*(Side(style='thin', color='B7B7B7') for _ in range(4)))


def ustaw_naglowek(ws, wiersz, kolumny):
    for i, nazwa in enumerate(kolumny, start=1):
        komorka = ws.cell(row=wiersz, column=i, value=nazwa)
        komorka.font = FONT_NAGLOWEK
        komorka.fill = FILL_NAGLOWEK
        komorka.alignment = WYROWNANIE_SRODEK
        komorka.border = OBRAMOWANIE_CIENKIE


def autoszerokosc(ws, min_szer=10, max_szer=42):
    for kolumna_komorki in ws.columns:
        dlugosc = 0
        litera = None
        for komorka in kolumna_komorki:
            if litera is None:
                litera = getattr(komorka, 'column_letter', None)
            if litera is None:
                continue
            if komorka.value is not None:
                dlugosc = max(dlugosc, len(str(komorka.value)))
        if litera:
            ws.column_dimensions[litera].width = max(min_szer, min(max_szer, dlugosc + 3))


def _pisz_tabele(ws, df, wiersz_start=1, formaty=None):
    ustaw_naglowek(ws, wiersz_start, list(df.columns))
    formaty = formaty or {}
    for i, wiersz in enumerate(df.itertuples(index=False), start=wiersz_start + 1):
        for j, (kolumna, wartosc) in enumerate(zip(df.columns, wiersz), start=1):
            if pd.isna(wartosc):
                wartosc = None
            elif isinstance(wartosc, float):
                wartosc = round(wartosc, formaty.get(kolumna, 3))
            komorka = ws.cell(row=i, column=j, value=wartosc)
            komorka.font = FONT_ZWYKLY
            komorka.border = OBRAMOWANIE_CIENKIE
    return wiersz_start + len(df)


def main(sciezka_csv=None, sciezka_xlsx=None):
    sciezka_csv = sciezka_csv or SCIEZKA_CSV_DOMYSLNA
    sciezka_xlsx = sciezka_xlsx or SCIEZKA_XLSX_DOMYSLNA

    df = pd.read_csv(sciezka_csv)
    df['szum'] = df['szum'].astype(bool)

    wb = Workbook()

    # ================= "Wyniki" - surowa tabela =================
    ws = wb.active
    ws.title = 'Wyniki'
    kolumny_wyniki = [c for c in [
        'lokalizacja', 'name', 'scenariusz', 'szum', 'perturb_k_pct', 'perturb_t1_pct', 'perturb_l_pct',
        'energia_kwh', 'przelaczenia', 'max_snieg_mm', 'max_lod_mm', 'max_hrt', 'min_hrt',
        'autotest_fit_ok', 'autotest_K', 'autotest_T1', 'autotest_T2', 'autotest_L',
        'blad_identyfikacji_K_pct', 'blad_identyfikacji_T1_pct', 'blad_identyfikacji_L_pct',
    ] if c in df.columns]
    ostatni = _pisz_tabele(ws, df[kolumny_wyniki].sort_values(['lokalizacja', 'name', 'scenariusz', 'szum']))
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(kolumny_wyniki)).coordinate[:-1]}{ostatni}'
    autoszerokosc(ws)

    # ================= "Odchylenie_energii" - % zmiany vs nominal (bez szumu) =================
    ws2 = wb.create_sheet('Odchylenie_energii')
    baseline = df[(df['scenariusz'] == 'nominal') & (~df['szum'])].set_index(['lokalizacja', 'name'])['energia_kwh']

    def _odchylenie(wiersz):
        klucz = (wiersz['lokalizacja'], wiersz['name'])
        base = baseline.get(klucz)
        if base is None or base == 0:
            return None
        return (wiersz['energia_kwh'] - base) / base * 100.0

    df_odch = df.copy()
    df_odch['odchylenie_energii_pct'] = df_odch.apply(_odchylenie, axis=1)
    kolumny_odch = ['lokalizacja', 'name', 'scenariusz', 'szum', 'energia_kwh', 'odchylenie_energii_pct']
    df_odch = df_odch[kolumny_odch].sort_values(['lokalizacja', 'name', 'scenariusz', 'szum'])
    ostatni2 = _pisz_tabele(ws2, df_odch)
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f'A1:F{ostatni2}'
    skala = ColorScaleRule(start_type='min', start_color='63BE7B', mid_type='num', mid_value=0, mid_color='FFEB84',
                            end_type='max', end_color='F8696B')
    ws2.conditional_formatting.add(f'F2:F{ostatni2}', skala)
    autoszerokosc(ws2)

    # ================= "Jakosc_autotestu" - błąd identyfikacji =================
    ws3 = wb.create_sheet('Jakosc_autotestu')
    df_auto = df[df['autotest_fit_ok'].notna()].copy()
    if not df_auto.empty:
        kolumny_auto = [c for c in [
            'lokalizacja', 'name', 'scenariusz', 'szum', 'autotest_fit_ok', 'autotest_r_squared',
            'blad_identyfikacji_K_pct', 'blad_identyfikacji_T1_pct', 'blad_identyfikacji_L_pct',
        ] if c in df_auto.columns]
        df_auto = df_auto[kolumny_auto].sort_values(['lokalizacja', 'name', 'scenariusz', 'szum'])
        ostatni3 = _pisz_tabele(ws3, df_auto)
        ws3.freeze_panes = 'A2'
        ws3.auto_filter.ref = f'A1:{ws3.cell(row=1, column=len(kolumny_auto)).coordinate[:-1]}{ostatni3}'
        for kolumna_nazwa in ('blad_identyfikacji_K_pct', 'blad_identyfikacji_T1_pct', 'blad_identyfikacji_L_pct'):
            if kolumna_nazwa in kolumny_auto:
                idx = kolumny_auto.index(kolumna_nazwa) + 1
                litera = ws3.cell(row=1, column=idx).coordinate[:-1]
                skala_auto = ColorScaleRule(start_type='num', start_value=-50, start_color='F8696B',
                                             mid_type='num', mid_value=0, mid_color='FFEB84',
                                             end_type='num', end_value=50, end_color='F8696B')
                ws3.conditional_formatting.add(f'{litera}2:{litera}{ostatni3}', skala_auto)
        autoszerokosc(ws3)
    else:
        ws3.cell(row=1, column=1, value='Brak algorytmów adaptacyjnych w wynikach (autotest_fit_ok puste).').font = FONT_ZWYKLY

    # ================= "Podsumowanie" - tekstowe wnioski =================
    ws4 = wb.create_sheet('Podsumowanie')
    ws4.cell(row=1, column=1, value='Wrażliwość na zaburzenie transmitancji + szum - podsumowanie').font = \
        Font(name=FONT_NAZWA, bold=True, size=14)

    wiersz = 3
    for lok in sorted(df['lokalizacja'].unique()):
        podz = df_odch[df_odch['lokalizacja'] == lok]
        podz_szum = podz[podz['szum']]
        podz_bez = podz[~podz['szum']]
        najbardziej_wrazliwy = podz_bez.loc[podz_bez['odchylenie_energii_pct'].abs().idxmax()] \
            if podz_bez['odchylenie_energii_pct'].notna().any() else None
        ws4.cell(row=wiersz, column=1, value=f'Lokalizacja: {lok}').font = FONT_POGRUBIONY
        wiersz += 1
        if najbardziej_wrazliwy is not None:
            ws4.cell(row=wiersz, column=1,
                     value=f"  Najbardziej wrażliwy na samo zaburzenie transmitancji (bez szumu): "
                           f"{najbardziej_wrazliwy['name']} / {najbardziej_wrazliwy['scenariusz']} "
                           f"({najbardziej_wrazliwy['odchylenie_energii_pct']:+.1f}% energii vs nominal)").font = FONT_ZWYKLY
            wiersz += 1
        if podz_szum['odchylenie_energii_pct'].notna().any():
            srednia_szum = podz_szum['odchylenie_energii_pct'].abs().mean()
            srednia_bez = podz_bez['odchylenie_energii_pct'].abs().mean()
            ws4.cell(row=wiersz, column=1,
                     value=f"  Średnie bezwzględne odchylenie energii: {srednia_bez:.1f}% bez szumu vs "
                           f"{srednia_szum:.1f}% z szumem (2.0°C std na HRT/CRT)").font = FONT_ZWYKLY
            wiersz += 1
        wiersz += 1

    if not df_auto.empty:
        ws4.cell(row=wiersz, column=1, value='Jakość autotestu (identyfikacja SOPDT)').font = FONT_POGRUBIONY
        wiersz += 1
        bez_szumu = df_auto[~df_auto['szum']]
        z_szumem = df_auto[df_auto['szum']]
        for etykieta, podz in (('bez szumu', bez_szumu), ('z szumem', z_szumem)):
            if podz.empty:
                continue
            sr_k = podz['blad_identyfikacji_K_pct'].abs().mean()
            sr_t1 = podz['blad_identyfikacji_T1_pct'].abs().mean()
            ws4.cell(row=wiersz, column=1,
                     value=f"  Średni bezwzględny błąd identyfikacji ({etykieta}): "
                           f"K={sr_k:.1f}%, T1={sr_t1:.1f}%").font = FONT_ZWYKLY
            wiersz += 1

    autoszerokosc(ws4)

    wb.save(sciezka_xlsx)
    print(f"Zapisano: {sciezka_xlsx}")


if __name__ == '__main__':
    main()
