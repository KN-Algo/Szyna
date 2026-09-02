# test_wrazliwosc_kroku_sterowania.py
#
# Sprawdza, jak CZĘSTO trzeba przeliczać decyzję sterowania (krok symulacji/
# sterowania dt - patrz KROK_SYMULACJI_S w test_wszystkie_rownolegle.py, obecny
# domyślny wybór to 10s) - dla wybranych (najlepszych) algorytmów, na WSZYSTKICH
# lokalizacjach pogodowych, porównuje energię i jakość regulacji (IAE/ISE/ITAE)
# przy kroku: 1s, 10s, 60s (1 min), 300s (5 min), 600s (10 min).
#
# Sens: dt=10s zostało wybrane wcześniej NA PODSTAWIE energii (różnica ~0.02%
# między dt=1 a dt=10 w jednym teście) - ten skrypt sprawdza to systematyczniej,
# na wielu lokalizacjach/algorytmach, i DODATKOWO patrzy na IAE/ISE/ITAE (jakość
# regulacji), nie tylko energię - rzadszy krok może dawać podobną energię, ale
# gorzej "trafiać" w cel chwila po chwili (wyższe IAE), co czysta energia by nie
# pokazała.
#
# Sterowanie (zmienne środowiskowe):
#   SZYNA_ALGORYTMY_KROK   - lista algorytmów do sprawdzenia (przecinki) -
#                             domyślnie 3 zwycięzcy wstępnego rankingu (2 lokalizacje,
#                             okno 30 dni, patrz AGENTS.md 2026-09-02): fuzzy_ryzyko_2v2_opad
#                             (#1 ogólnie, rodzina "funkcja ryzyka"), fuzzy_normy_2v2
#                             (#8, najlepszy z rodziny "progi normy" - inne stroalne
#                             stałe niż #1), nauka_kary_opad (najlepszy z rodziny
#                             "uczenie z kar" - jeszcze inne stroalne stałe) - CELOWO
#                             zróżnicowani (nie literalne top-3 wg samej energii, bo
#                             te są prawie identyczne warianty fuzzy_ryzyko_* strojące
#                             TE SAME stałe RISK_* - nieinformatywne dla grid-search).
#   SZYNA_KROKI_S           - lista kroków [s] do przetestowania (przecinki),
#                             domyślnie "1,10,60,300,600"
#   SZYNA_LOKALIZACJE       - jak w test_wszystkie_rownolegle.py (domyślnie wszystkie)
#   SZYNA_MAX_DNI           - jak w test_wszystkie_rownolegle.py (domyślnie 30 -
#                             skrócone dla szybkości, to analiza porównawcza,
#                             nie ostateczny wynik na produkcję)
#   SZYNA_LICZBA_WATKOW     - jak reszta projektu
#   SZYNA_FOLDER_WYNIKOW_KROK - folder wyników (domyślnie wyniki/wrazliwosc_kroku)
#   SZYNA_WZNOW             - wznawianie (domyślnie 1)
#
# Uruchomienie: python test_wrazliwosc_kroku_sterowania.py

import os
import sys
import time
import traceback
from concurrent.futures import ProcessPoolExecutor, as_completed

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_POGODA = os.path.join(BASE_DIR, "Pogoda_pomiary_15_minut")
FOLDER_WYNIKOW = os.environ.get(
    'SZYNA_FOLDER_WYNIKOW_KROK', os.path.join(BASE_DIR, "wyniki", "wrazliwosc_kroku"))
os.makedirs(FOLDER_WYNIKOW, exist_ok=True)

_alg_env = os.environ.get('SZYNA_ALGORYTMY_KROK', 'fuzzy_ryzyko_2v2_opad,fuzzy_normy_2v2,nauka_kary_opad')
ALGORYTMY_KROK = [a.strip() for a in _alg_env.split(',') if a.strip()]

_kroki_env = os.environ.get('SZYNA_KROKI_S', '1,10,60,300,600')
KROKI_S = [float(k.strip()) for k in _kroki_env.split(',') if k.strip()]

_lok_env = os.environ.get('SZYNA_LOKALIZACJE')
LOKALIZACJE_FILTR = {l.strip() for l in _lok_env.split(',') if l.strip()} if _lok_env else None

_max_dni_env = os.environ.get('SZYNA_MAX_DNI', '30')
MAX_DNI = int(_max_dni_env) if _max_dni_env else None

MAX_SWITCHES_PER_DAY = 100
WZNAWIAJ_PRZERWANE = os.environ.get('SZYNA_WZNOW', '1') != '0'

_watkow_env = os.environ.get('SZYNA_LICZBA_WATKOW')
LICZBA_WATKOW_NADPISANIE = int(_watkow_env) if _watkow_env else None

FONT_NAZWA = 'Arial'
FONT_NAGLOWEK = Font(name=FONT_NAZWA, bold=True, color='FFFFFF', size=11)
FILL_NAGLOWEK = PatternFill('solid', fgColor='1F4E78')
FONT_ZWYKLY = Font(name=FONT_NAZWA, size=10)
OBRAMOWANIE = Border(*(Side(style='thin', color='B7B7B7') for _ in range(4)))


def wykryj_liczbe_watkow():
    if LICZBA_WATKOW_NADPISANIE:
        return LICZBA_WATKOW_NADPISANIE
    for zmienna in ('SLURM_CPUS_PER_TASK', 'SLURM_JOB_CPUS_PER_NODE', 'PBS_NP', 'NSLOTS', 'LSB_DJOB_NUMPROC'):
        wartosc = os.environ.get(zmienna)
        if wartosc:
            try:
                liczba = int(str(wartosc).split('(')[0].split(',')[0])
                if liczba > 0:
                    return liczba
            except ValueError:
                continue
    try:
        return len(os.sched_getaffinity(0))
    except AttributeError:
        return os.cpu_count() or 1


def znajdz_pliki_pogodowe():
    pliki = {
        os.path.splitext(nazwa)[0]: os.path.join(FOLDER_POGODA, nazwa)
        for nazwa in sorted(os.listdir(FOLDER_POGODA)) if nazwa.endswith('.csv')
    }
    if LOKALIZACJE_FILTR is not None:
        pliki = {k: v for k, v in pliki.items() if k in LOKALIZACJE_FILTR}
    return pliki


def przetworz_zadanie(nazwa_lokalizacji, sciezka_csv, nazwa_algorytmu, krok_s):
    try:
        sys.path.insert(0, os.path.join(BASE_DIR, 'Algorytmy'))
        import symulacja_fizyczna as fiz
        from rejestr_algorytmow import stworz_kontroler, podlega_bezpiecznikowi

        zakres_dat = fiz.wybierz_najzimniejsze_okno(sciezka_csv, MAX_DNI) if MAX_DNI else None
        dt = krok_s
        df_1s = fiz.wczytaj_pogode_1s(sciezka_csv, zakres_dat=zakres_dat, dt=dt)
        A_wd, B_wd, C_wd, D_wd, A_hd, B_hd, C_hd, D_hd, punkty_opoznienia = fiz.przygotuj_modele_stanowe(dt)
        at_array = df_1s['temperatura_powietrza_C'].to_numpy()
        hrt_weather_all = fiz.wylicz_skladowa_pogodowa(at_array, A_wd, B_wd, C_wd, D_wd, dt)

        kontroler_normy, metoda_normy = stworz_kontroler('algorytm_z_normy', max_switches_per_day=MAX_SWITCHES_PER_DAY)
        df_normy, stats_normy, snow_ref, power_ref = fiz.uruchom_kontroler(
            'algorytm_z_normy', kontroler_normy, metoda_normy, df_1s, hrt_weather_all,
            A_hd, B_hd, C_hd, D_hd, punkty_opoznienia, dt=dt, print_progress=False,
        )

        czy_bezpiecznik = podlega_bezpiecznikowi(nazwa_algorytmu)
        kontroler, metoda = stworz_kontroler(nazwa_algorytmu, max_switches_per_day=MAX_SWITCHES_PER_DAY)
        _, stats, _, _ = fiz.uruchom_kontroler(
            nazwa_algorytmu, kontroler, metoda, df_1s, hrt_weather_all,
            A_hd, B_hd, C_hd, D_hd, punkty_opoznienia, dt=dt,
            snow_reference_mm=snow_ref if czy_bezpiecznik else None,
            power_reference_pct=power_ref if czy_bezpiecznik else None,
            print_progress=False,
        )

        stats = dict(stats)
        stats['lokalizacja'] = nazwa_lokalizacji
        stats['krok_s'] = krok_s
        return nazwa_lokalizacji, nazwa_algorytmu, krok_s, stats, None
    except Exception:
        return nazwa_lokalizacji, nazwa_algorytmu, krok_s, None, traceback.format_exc()


def main():
    liczba_watkow = wykryj_liczbe_watkow()
    pliki_pogodowe = znajdz_pliki_pogodowe()
    print(f"Wrażliwość na krok sterowania: {len(ALGORYTMY_KROK)} algorytmów ({', '.join(ALGORYTMY_KROK)}), "
          f"{len(KROKI_S)} kroków ({KROKI_S}), {len(pliki_pogodowe)} lokalizacji, "
          f"okno {MAX_DNI if MAX_DNI else 'pełne'} dni, {liczba_watkow} procesów.\n")

    zadania = [
        (lok, sciezka, alg, krok)
        for lok, sciezka in pliki_pogodowe.items()
        for alg in ALGORYTMY_KROK
        for krok in KROKI_S
    ]
    liczba_zadan_ogolem = len(zadania)
    print(f"Łącznie {liczba_zadan_ogolem} zadań.\n")

    wyniki = []
    sciezka_zbiorczy = os.path.join(FOLDER_WYNIKOW, "WRAZLIWOSC_KROKU_ZBIORCZY.csv")
    if WZNAWIAJ_PRZERWANE and os.path.exists(sciezka_zbiorczy):
        try:
            df_poprzedni = pd.read_csv(sciezka_zbiorczy)
            wyniki = df_poprzedni.to_dict('records')
            gotowe = {(w['lokalizacja'], w['name'], float(w['krok_s'])) for w in wyniki}
            zadania = [z for z in zadania if (z[0], z[2], z[3]) not in gotowe]
            print(f"WZNOWIENIE: {len(gotowe)} gotowych - liczą się tylko brakujące {len(zadania)}/{liczba_zadan_ogolem}.\n")
        except Exception:
            wyniki = []

    if not zadania:
        print("Wszystkie zadania już policzone.")
    else:
        bledy = []
        t0 = time.time()
        with ProcessPoolExecutor(max_workers=liczba_watkow) as executor:
            futures = {executor.submit(przetworz_zadanie, lok, sciezka, alg, krok): (lok, alg, krok)
                       for lok, sciezka, alg, krok in zadania}
            zakonczone = 0
            for future in as_completed(futures):
                lok, alg, krok, stats, blad = future.result()
                zakonczone += 1
                elapsed_min = (time.time() - t0) / 60.0
                etykieta = f"{lok}/{alg}/dt={krok:g}s"
                if blad is not None:
                    bledy.append((etykieta, blad))
                    print(f"[{zakonczone}/{liczba_zadan_ogolem}] BŁĄD {etykieta}:\n{blad}")
                else:
                    wyniki.append(stats)
                    print(f"[{zakonczone}/{liczba_zadan_ogolem}] OK {etykieta} energia={stats['energia_kwh']:.1f} kWh "
                          f"(upłynęło {elapsed_min:.1f} min)")
                pd.DataFrame(wyniki).to_csv(sciezka_zbiorczy, index=False)
        print(f"\nZakończono. Sukcesy: {len(wyniki)}/{liczba_zadan_ogolem}. Błędy: {len(bledy)}.")

    if not wyniki:
        print("Brak wyników.")
        return

    df = pd.DataFrame(wyniki)
    zbuduj_excel(df)


def zbuduj_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Wyniki'
    kolumny = [c for c in ['lokalizacja', 'name', 'krok_s', 'energia_kwh', 'przelaczenia',
                            'max_snieg_mm', 'max_hrt', 'min_hrt', 'iae', 'ise', 'itae'] if c in df.columns]
    for j, nazwa in enumerate(kolumny, start=1):
        c = ws.cell(row=1, column=j, value=nazwa)
        c.font = FONT_NAGLOWEK
        c.fill = FILL_NAGLOWEK
        c.alignment = Alignment(horizontal='center')
    for i, wiersz in enumerate(df[kolumny].itertuples(index=False), start=2):
        for j, wartosc in enumerate(wiersz, start=1):
            v = round(wartosc, 3) if isinstance(wartosc, float) and pd.notna(wartosc) else \
                (None if isinstance(wartosc, float) and pd.isna(wartosc) else wartosc)
            cell = ws.cell(row=i, column=j, value=v)
            cell.font = FONT_ZWYKLY
            cell.border = OBRAMOWANIE
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(kolumny)).coordinate[:-1]}{len(df) + 1}'

    ws2 = wb.create_sheet('Srednie_wg_kroku')
    agg = df.groupby(['name', 'krok_s']).agg(
        energia_srednia=('energia_kwh', 'mean'),
        iae_srednia=('iae', 'mean') if 'iae' in df.columns else ('energia_kwh', 'mean'),
    ).reset_index()

    naglowki2 = ['Algorytm', 'Krok [s]', 'Średnia energia (kWh)', 'Średnie IAE (°C·s)']
    for j, nazwa in enumerate(naglowki2, start=1):
        c = ws2.cell(row=1, column=j, value=nazwa)
        c.font = FONT_NAGLOWEK
        c.fill = FILL_NAGLOWEK
    for i, wiersz in enumerate(agg.itertuples(index=False), start=2):
        ws2.cell(row=i, column=1, value=wiersz.name)
        ws2.cell(row=i, column=2, value=wiersz.krok_s)
        ws2.cell(row=i, column=3, value=round(wiersz.energia_srednia, 2))
        iae_val = getattr(wiersz, 'iae_srednia', None)
        ws2.cell(row=i, column=4, value=round(iae_val, 1) if pd.notna(iae_val) else None)

    # Wykres: energia vs krok sterowania, jedna linia na algorytm.
    ostatni2 = len(agg) + 1
    chart = LineChart()
    chart.title = 'Średnia energia vs krok sterowania'
    chart.x_axis.title = 'Krok sterowania [s]'
    chart.y_axis.title = 'Energia (kWh)'
    chart.width = 24
    chart.height = 12
    wiersz_pomoc = ostatni2 + 3
    for algorytm in sorted(agg['name'].unique()):
        podz = agg[agg['name'] == algorytm].sort_values('krok_s')
        ws2.cell(row=wiersz_pomoc, column=6, value=algorytm).font = Font(name=FONT_NAZWA, bold=True)
        for k, (_, w) in enumerate(podz.iterrows(), start=1):
            ws2.cell(row=wiersz_pomoc + k, column=5, value=w['krok_s'])
            ws2.cell(row=wiersz_pomoc + k, column=6, value=round(w['energia_srednia'], 2))
        dane_serii = Reference(ws2, min_col=6, min_row=wiersz_pomoc, max_row=wiersz_pomoc + len(podz))
        kategorie = Reference(ws2, min_col=5, min_row=wiersz_pomoc + 1, max_row=wiersz_pomoc + len(podz))
        chart.add_data(dane_serii, titles_from_data=True)
        chart.set_categories(kategorie)
        wiersz_pomoc += len(podz) + 2
    ws2.add_chart(chart, 'H2')
    for kol in ('A', 'B', 'C', 'D'):
        ws2.column_dimensions[kol].width = 22

    autoszerokosc(ws)

    sciezka_xlsx = os.path.join(FOLDER_WYNIKOW, "Podsumowanie_kroku_sterowania.xlsx")
    wb.save(sciezka_xlsx)
    print(f"Zapisano: {sciezka_xlsx}")


def autoszerokosc(ws, min_szer=10, max_szer=42):
    for kolumna_komorki in ws.columns:
        dlugosc = 0
        litera = None
        for komorka in kolumna_komorki:
            if litera is None:
                litera = getattr(komorka, 'column_letter', None)
            if litera is None:
                continue
            if komorka.value is not None:
                dlugosc = max(dlugosc, len(str(komorka.value)))
        if litera:
            ws.column_dimensions[litera].width = max(min_szer, min(max_szer, dlugosc + 3))


if __name__ == '__main__':
    main()
