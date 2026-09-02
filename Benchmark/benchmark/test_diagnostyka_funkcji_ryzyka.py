# test_diagnostyka_funkcji_ryzyka.py
#
# Diagnostyka WIZUALNA: jak funkcja ryzyka (i porównywalne algorytmy) NAPRAWDĘ
# liczy wartości krok po kroku - AT/CRT/HRT (fizyka) razem z target_temperature
# i need_heat (decyzja algorytmu) na jednym wykresie, żeby zobaczyć NA OKO, czy
# kaskada priorytetów (patrz Algorytmy/funkcja_ryzyka_wspolne.py) reaguje
# sensownie na prawdziwą pogodę - a nie tylko wierzyć liczbom zbiorczym
# (energia/przełączenia) w PRZEGLAD_ZBIORCZY.csv.
#
# Wykorzystuje kolumny Target_temperature/Need_heat, które
# symulacja_fizyczna.uruchom_kontroler teraz zawsze zapisuje do df_hist (NaN
# dla algorytmów bez jawnego, ciągłego celu - patrz IAE/ISE/ITAE) - żaden
# dodatkowy hak, korzysta z tego samego mechanizmu.
#
# Domyślnie diagnozuje risk_function_pid (flagowy wariant funkcji ryzyka) na
# Abisko (najbardziej zróżnicowana pogoda ze wszystkich lokalizacji - dobra
# szansa zobaczyć WSZYSTKIE 4 priorytety kaskady w akcji), okno 14 dni (na
# tyle krótkie, żeby wykres był czytelny, na tyle długie, żeby zobaczyć kilka
# epizodów). Można podać kilka algorytmów naraz (SZYNA_ALGORYTMY_DIAG) - każdy
# dostaje własną zakładkę do porównania.
#
# Sterowanie (zmienne środowiskowe):
#   SZYNA_LOKALIZACJA_DIAG   - lokalizacja (domyślnie abisko_60min_2024)
#   SZYNA_ALGORYTMY_DIAG     - lista algorytmów, przecinki (domyślnie "risk_function_pid")
#   SZYNA_MAX_DNI_DIAG       - okno dni, najzimniejszy wycinek (domyślnie 14)
#   SZYNA_KROK_S             - krok symulacji [s] (domyślnie 10.0)
#   SZYNA_FOLDER_WYNIKOW_DIAG - folder wyników (domyślnie wyniki/diagnostyka_funkcji_ryzyka)
#
# Uruchomienie: python test_diagnostyka_funkcji_ryzyka.py

import os
import sys

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_POGODA = os.path.join(BASE_DIR, "Pogoda_pomiary_15_minut")
FOLDER_WYNIKOW = os.environ.get(
    'SZYNA_FOLDER_WYNIKOW_DIAG', os.path.join(BASE_DIR, "wyniki", "diagnostyka_funkcji_ryzyka"))
os.makedirs(FOLDER_WYNIKOW, exist_ok=True)

LOKALIZACJA = os.environ.get('SZYNA_LOKALIZACJA_DIAG', 'abisko_60min_2024')
_alg_env = os.environ.get('SZYNA_ALGORYTMY_DIAG', 'risk_function_pid')
ALGORYTMY_DIAG = [a.strip() for a in _alg_env.split(',') if a.strip()]
MAX_DNI = int(os.environ.get('SZYNA_MAX_DNI_DIAG', '14'))
KROK_SYMULACJI_S = float(os.environ.get('SZYNA_KROK_S', '10.0'))
MAX_SWITCHES_PER_DAY = 100

FONT_NAZWA = 'Arial'


def uruchom_jeden(nazwa_algorytmu):
    sys.path.insert(0, os.path.join(BASE_DIR, 'Algorytmy'))
    import symulacja_fizyczna as fiz
    from rejestr_algorytmow import stworz_kontroler, podlega_bezpiecznikowi

    sciezka_csv = os.path.join(FOLDER_POGODA, f"{LOKALIZACJA}.csv")
    zakres_dat = fiz.wybierz_najzimniejsze_okno(sciezka_csv, MAX_DNI)
    dt = KROK_SYMULACJI_S
    df_1s = fiz.wczytaj_pogode_1s(sciezka_csv, zakres_dat=zakres_dat, dt=dt)
    A_wd, B_wd, C_wd, D_wd, A_hd, B_hd, C_hd, D_hd, punkty_opoznienia = fiz.przygotuj_modele_stanowe(dt)
    at_array = df_1s['temperatura_powietrza_C'].to_numpy()
    hrt_weather_all = fiz.wylicz_skladowa_pogodowa(at_array, A_wd, B_wd, C_wd, D_wd, dt)

    kontroler_normy, metoda_normy = stworz_kontroler('algorytm_z_normy', max_switches_per_day=MAX_SWITCHES_PER_DAY)
    df_normy, stats_normy, snow_ref, power_ref = fiz.uruchom_kontroler(
        'algorytm_z_normy', kontroler_normy, metoda_normy, df_1s, hrt_weather_all,
        A_hd, B_hd, C_hd, D_hd, punkty_opoznienia, dt=dt, print_progress=False,
    )

    if nazwa_algorytmu == 'algorytm_z_normy':
        df_hist, stats = df_normy, stats_normy
    else:
        czy_bezpiecznik = podlega_bezpiecznikowi(nazwa_algorytmu)
        kontroler, metoda = stworz_kontroler(nazwa_algorytmu, max_switches_per_day=MAX_SWITCHES_PER_DAY)
        df_hist, stats, _, _ = fiz.uruchom_kontroler(
            nazwa_algorytmu, kontroler, metoda, df_1s, hrt_weather_all,
            A_hd, B_hd, C_hd, D_hd, punkty_opoznienia, dt=dt,
            snow_reference_mm=snow_ref if czy_bezpiecznik else None,
            power_reference_pct=power_ref if czy_bezpiecznik else None,
            print_progress=False,
        )

    # Downsampling do 15 min - wystarczająco gęste, żeby zobaczyć dynamikę
    # kaskady, wystarczająco rzadkie, żeby wykres Excela pozostał czytelny/lekki.
    df_zapis = fiz.przygotuj_do_zapisu(df_hist, 900)
    return df_zapis, stats


def _dodaj_arkusz(wb, nazwa_algorytmu, df):
    ws = wb.create_sheet(nazwa_algorytmu[:31])
    kolumny = ['Timestamp', 'AT', 'CRT', 'HRT', 'Target_temperature', 'Need_heat', 'Moc_procent', 'Snieg_mm']
    kolumny = [k for k in kolumny if k in df.columns]
    for j, nazwa in enumerate(kolumny, start=1):
        ws.cell(row=1, column=j, value=nazwa).font = Font(name=FONT_NAZWA, bold=True)
    for i, wiersz in enumerate(df[kolumny].itertuples(index=False), start=2):
        for j, wartosc in enumerate(wiersz, start=1):
            if isinstance(wartosc, float):
                wartosc = round(wartosc, 3)
            ws.cell(row=i, column=j, value=str(wartosc) if kolumny[j - 1] == 'Timestamp' else wartosc)

    ostatni = len(df) + 1
    ma_target = 'Target_temperature' in kolumny

    chart_temp = LineChart()
    chart_temp.title = f'{nazwa_algorytmu} - temperatury i cel'
    chart_temp.y_axis.title = 'Temperatura (°C)'
    chart_temp.x_axis.title = 'Krok (co 15 min)'
    chart_temp.width = 26
    chart_temp.height = 12
    for nazwa_kol in ('AT', 'CRT', 'HRT', 'Target_temperature'):
        if nazwa_kol not in kolumny:
            continue
        idx = kolumny.index(nazwa_kol) + 1
        dane = Reference(ws, min_col=idx, min_row=1, max_row=ostatni)
        chart_temp.add_data(dane, titles_from_data=True)
    ws.add_chart(chart_temp, f'J2')

    if 'Moc_procent' in kolumny:
        chart_moc = LineChart()
        chart_moc.title = f'{nazwa_algorytmu} - moc grzania (%)'
        chart_moc.y_axis.title = 'Moc (%)'
        chart_moc.x_axis.title = 'Krok (co 15 min)'
        chart_moc.width = 26
        chart_moc.height = 10
        idx = kolumny.index('Moc_procent') + 1
        dane = Reference(ws, min_col=idx, min_row=1, max_row=ostatni)
        chart_moc.add_data(dane, titles_from_data=True)
        ws.add_chart(chart_moc, 'J28')

    return ma_target


def main():
    print(f"Diagnostyka funkcji ryzyka: lokalizacja={LOKALIZACJA}, {MAX_DNI} dni, krok={KROK_SYMULACJI_S:g}s")
    print(f"Algorytmy: {', '.join(ALGORYTMY_DIAG)}\n")

    wb = Workbook()
    wb.remove(wb.active)

    podsumowania = []
    for nazwa_algorytmu in ALGORYTMY_DIAG:
        print(f"Liczę {nazwa_algorytmu}...")
        df_zapis, stats = uruchom_jeden(nazwa_algorytmu)
        sciezka_csv = os.path.join(FOLDER_WYNIKOW, f"{LOKALIZACJA}_{nazwa_algorytmu}_diagnostyka.csv")
        df_zapis.to_csv(sciezka_csv, index=False)
        ma_target = _dodaj_arkusz(wb, nazwa_algorytmu, df_zapis)
        podsumowania.append(dict(
            algorytm=nazwa_algorytmu, energia_kwh=stats['energia_kwh'],
            iae=stats.get('iae'), ise=stats.get('ise'), itae=stats.get('itae'),
            ma_cel_ciagly=ma_target,
        ))
        print(f"  -> energia={stats['energia_kwh']:.1f} kWh, IAE={stats.get('iae')}, "
              f"zapisano {sciezka_csv}")

    ws_sum = wb.create_sheet('Podsumowanie', 0)
    naglowki = ['Algorytm', 'Energia (kWh)', 'IAE (°C·s)', 'ISE (°C²·s)', 'ITAE (°C·s²)', 'Ma jawny cel ciągły']
    for j, nazwa in enumerate(naglowki, start=1):
        ws_sum.cell(row=1, column=j, value=nazwa).font = Font(name=FONT_NAZWA, bold=True)
    for i, wpis in enumerate(podsumowania, start=2):
        wartosci = [wpis['algorytm'], round(wpis['energia_kwh'], 1),
                    round(wpis['iae'], 1) if wpis['iae'] is not None else None,
                    round(wpis['ise'], 1) if wpis['ise'] is not None else None,
                    round(wpis['itae'], 1) if wpis['itae'] is not None else None,
                    'Tak' if wpis['ma_cel_ciagly'] else 'Nie']
        for j, wartosc in enumerate(wartosci, start=1):
            ws_sum.cell(row=i, column=j, value=wartosc)

    sciezka_xlsx = os.path.join(FOLDER_WYNIKOW, f"Diagnostyka_{LOKALIZACJA}.xlsx")
    wb.save(sciezka_xlsx)
    print(f"\nGotowe. Excel: {sciezka_xlsx}")


if __name__ == '__main__':
    main()
