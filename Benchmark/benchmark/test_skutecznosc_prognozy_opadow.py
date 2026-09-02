# test_skutecznosc_prognozy_opadow.py
#
# Waliduje SKUTECZNOŚĆ modułu przewidywanie_opadow.py - dokładnie tej samej
# klasy `przewidywanie_opadow`, która jest realnie używana w kontrolerach
# *_opad.py (patrz Algorytmy/funkcja_ryzyka_wspolne.py:
# KontrolerRyzykaOpadBazowy._prognoza_intensywnosci_opadu) - na WSZYSTKICH
# dostępnych próbkach pogodowych (Pogoda_pomiary_15_minut/*.csv).
#
# Dla każdego pliku porównuje prognozę algorytmu z rzeczywistym opadem
# (ground truth wyliczone z opad_mm/temperatura_powietrza_C/wiatr_m_s - te
# same progi co w oryginalnym module testującym w przewidywanie_opadow.py),
# licząc:
#   - skuteczność osłony (% godzin opadu zimowego, które algorytm wykrył)
#   - trafność alarmu / precyzję (% podniesionych alarmów, które były trafne)
#   - dokładność dokładnego poziomu intensywności (0-3)
#   - dokładność w funkcji horyzontu prognozy (krok 1..8 naprzód)
#   - rozbicie fałszywych alarmów: za wcześnie / za późno / czysto fałszywe
#
# Kroki źródłowe plików RÓŻNIĄ SIĘ (15 min dla suwalki/wroclaw, 60 min dla
# pozostałych 40 lokalizacji) - krok jest AUTOMATYCZNIE wykrywany z odstępów
# czasowych w każdym pliku (patrz _wykryj_krok_probkowania_h) i użyty do
# przeliczenia liczby próbek na godziny; sam algorytm i progi oceny
# (0.35 mm / 1.0 mm) są celowo NIE zmieniane względem oryginału w
# przewidywanie_opadow.py - to jest test istniejącego, wdrożonego modułu
# "as-is" na każdym pliku w jego naturalnej rozdzielczości, nie osobno
# dostrojony wariant per-lokalizacja.
#
# To liczy się LOKALNIE (bez klastra) w kilka minut - pasek postępu w
# terminalu pokazuje % ukończenia całości (ważony liczbą ocenianych próbek).
#
# Wynik: Excel wyniki/Podsumowanie_prognozy_opadow.xlsx:
#   - "Wyniki_lokalizacje"  - jeden wiersz na plik pogodowy (43 wiersze)
#   - "Podsumowanie_ogolne" - statystyki zbiorcze (globalne pooled + średnia
#                              po lokalizacjach) + wykres dokładności wg horyzontu
#   - "Definicje_poziomow"  - progi definiujące poziomy intensywności 0-3
#
# Uruchomienie (z katalogu Benchmark/benchmark):
#   python test_skutecznosc_prognozy_opadow.py

import os
import re
import sys
import time
import glob

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import LineChart, Reference

from przewidywanie_opadow import przewidywanie_opadow as PrzewidywanieOpadow

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_POGODA = os.path.join(BASE_DIR, 'Pogoda_pomiary_15_minut')
FOLDER_WYNIKOW = os.path.join(BASE_DIR, 'wyniki')
SCIEZKA_XLSX = os.path.join(FOLDER_WYNIKOW, 'Podsumowanie_prognozy_opadow.xlsx')

# Te same wartości co w przewidywanie_opadow.algorytm_opadu (oryginalny,
# wdrożony test modułu) - celowo NIE zmieniane.
LOOKBACK = 4
HORYZONT = 8

INTENSITY_DEFS = {
    0: ("Brak opadu", "0.00 mm / krok"),
    1: ("Słaby opad", "0.01 - 0.35 mm / krok (lekki śnieg / mżawka marznąca)"),
    2: ("Średni opad", "0.36 - 1.00 mm / krok (umiarkowany śnieg / deszcz marznący), wiatr < 6 m/s"),
    3: ("Mocny opad / Nawałnica", "> 1.00 mm / krok, LUB średni opad przy wietrze >= 6 m/s"),
}

# ==========================================================================
# STYL (spójny z generuj_excel_podsumowanie.py)
# ==========================================================================
FONT_NAZWA = 'Arial'
KOLOR_NAGLOWEK_BG = '1F4E78'
KOLOR_NAGLOWEK_FG = 'FFFFFF'

FONT_NAGLOWEK = Font(name=FONT_NAZWA, bold=True, color=KOLOR_NAGLOWEK_FG, size=11)
FILL_NAGLOWEK = PatternFill('solid', fgColor=KOLOR_NAGLOWEK_BG)
FONT_ZWYKLY = Font(name=FONT_NAZWA, size=10)
FONT_POGRUBIONY = Font(name=FONT_NAZWA, bold=True, size=10)
WYROWNANIE_SRODEK = Alignment(horizontal='center', vertical='center')
OBRAMOWANIE_CIENKIE = Border(*(Side(style='thin', color='B7B7B7') for _ in range(4)))


def ustaw_naglowek(ws, wiersz, kolumny):
    for i, nazwa in enumerate(kolumny, start=1):
        komorka = ws.cell(row=wiersz, column=i, value=nazwa)
        komorka.font = FONT_NAGLOWEK
        komorka.fill = FILL_NAGLOWEK
        komorka.alignment = WYROWNANIE_SRODEK
        komorka.border = OBRAMOWANIE_CIENKIE


def autoszerokosc(ws, min_szer=10, max_szer=42):
    for kolumna_komorki in ws.columns:
        dlugosc = 0
        litera = None
        for komorka in kolumna_komorki:
            if litera is None:
                litera = getattr(komorka, 'column_letter', None)
            if litera is None:
                continue
            if komorka.value is not None:
                dlugosc = max(dlugosc, len(str(komorka.value)))
        if litera:
            ws.column_dimensions[litera].width = max(min_szer, min(max_szer, dlugosc + 3))


# ==========================================================================
# WCZYTYWANIE I OCENA POJEDYNCZEGO PLIKU
# ==========================================================================
def _nazwa_lokalizacji(plik):
    baza = os.path.splitext(os.path.basename(plik))[0]
    m = re.match(r'^(.+)_(\d+)min_(\d{4})$', baza)
    if m:
        return m.group(1), int(m.group(2)), int(m.group(3))
    return baza, None, None


def _wykryj_krok_probkowania_h(df):
    delty = df['Timestamp'].diff().dropna()
    if delty.empty:
        return 0.25
    mediana_s = delty.dt.total_seconds().median()
    return float(mediana_s) / 3600.0


def wczytaj_plik(plik):
    df = pd.read_csv(plik)
    if 'data_czas' in df.columns:
        df.rename(columns={'data_czas': 'Timestamp'}, inplace=True)
    df['Timestamp'] = pd.to_datetime(df['Timestamp'])
    df.sort_values('Timestamp', inplace=True)
    df.reset_index(drop=True, inplace=True)
    if 'wiatr_m_s' not in df.columns:
        df['wiatr_m_s'] = 3.0
    return df


def wylicz_ground_truth(df):
    # Te same progi co w przewidywanie_opadow.algorytm_opadu.
    opad = df['opad_mm'].to_numpy(dtype=float)
    temp = df['temperatura_powietrza_C'].to_numpy(dtype=float)
    wiatr = df['wiatr_m_s'].to_numpy(dtype=float)

    gt = np.zeros(len(df), dtype=int)
    zimowy = (opad > 0.0001) & (temp <= 1.0)
    gt[zimowy & (opad <= 0.35)] = 1
    srednio_lub_mocno = zimowy & (opad > 0.35) & (opad <= 1.0)
    gt[srednio_lub_mocno & (wiatr < 6.0)] = 2
    gt[srednio_lub_mocno & (wiatr >= 6.0)] = 3
    gt[zimowy & (opad > 1.0)] = 3
    return gt


def oceniaj_plik(plik, callback_postepu=None):
    df = wczytaj_plik(plik)
    krok_h = _wykryj_krok_probkowania_h(df)
    gt_all = wylicz_ground_truth(df)

    forecaster = PrzewidywanieOpadow(persistence_steps=1)

    opad_arr = df['opad_mm'].to_numpy(dtype=float)
    temp_arr = df['temperatura_powietrza_C'].to_numpy(dtype=float)
    dp_arr = df['punkt_rosy_C'].to_numpy(dtype=float)
    wiatr_arr = df['wiatr_m_s'].to_numpy(dtype=float)

    eval_indices = list(range(LOOKBACK - 1, len(df) - HORYZONT))
    n_eval = len(eval_indices)
    if n_eval <= 0:
        raise ValueError(f"Plik {plik} jest za krótki do oceny (LOOKBACK+HORYZONT={LOOKBACK + HORYZONT} próbek).")

    pred_matrix = np.zeros((n_eval, HORYZONT), dtype=int)
    for pos, i in enumerate(eval_indices):
        past_precip = opad_arr[i - LOOKBACK + 1: i + 1]
        future_t = temp_arr[i + 1: i + 1 + HORYZONT]
        current_dp = dp_arr[i]
        current_wind = wiatr_arr[i]
        pred_matrix[pos] = forecaster.predict_winter_precipitation(
            past_precip, future_t, current_dp, current_wind
        )
        if callback_postepu and (pos % 200 == 0 or pos == n_eval - 1):
            callback_postepu(pos + 1, n_eval)

    # --- Macierz "stanów" (czy grzać TERAZ, krok +1) + rozbicie fałszywych alarmów ---
    tp = fn = tn = 0
    wczesnie = pozno = falszywe = 0
    for pos, i in enumerate(eval_indices):
        target_idx = i + 1
        actual = gt_all[target_idx]
        pred = pred_matrix[pos, 0]
        if pred > 0 and actual > 0:
            tp += 1
        elif pred == 0 and actual > 0:
            fn += 1
        elif pred == 0 and actual == 0:
            tn += 1
        else:
            future_gt = gt_all[target_idx + 1: min(target_idx + 9, len(gt_all))]
            past_gt = gt_all[max(0, target_idx - 8): target_idx]
            if np.any(future_gt > 0):
                wczesnie += 1
            elif np.any(past_gt > 0):
                pozno += 1
            else:
                falszywe += 1

    # --- Dokładność dokładnego poziomu (0-3) na krok +1 ---
    poziom_total = np.zeros(4, dtype=int)
    poziom_trafienia = np.zeros(4, dtype=int)
    for pos, i in enumerate(eval_indices):
        g = int(gt_all[i + 1])
        p = int(pred_matrix[pos, 0])
        poziom_total[g] += 1
        if p == g:
            poziom_trafienia[g] += 1

    # --- Dokładność w funkcji horyzontu (krok 1..HORYZONT naprzód) ---
    horyzont_trafienia = np.zeros(HORYZONT, dtype=int)
    for h in range(HORYZONT):
        cele = np.array([gt_all[i + 1 + h] for i in eval_indices])
        horyzont_trafienia[h] = int(np.sum(pred_matrix[:, h] == cele))

    godziny_analizy = n_eval * krok_h
    godziny_opadu = float(np.sum(gt_all[eval_indices] > 0)) * krok_h
    godziny_wykryte = tp * krok_h
    godziny_przegapione = fn * krok_h
    godziny_puste = (wczesnie + pozno + falszywe) * krok_h

    skutecznosc_oslony = (godziny_wykryte / godziny_opadu * 100) if godziny_opadu > 0 else np.nan
    wszystkie_alarmy = tp + wczesnie + pozno + falszywe
    trafnosc_alarmu = (tp / wszystkie_alarmy * 100) if wszystkie_alarmy > 0 else np.nan
    dokladnosc_poziomow = float(np.sum(poziom_trafienia)) / n_eval * 100

    return {
        'krok_h': krok_h,
        'n_eval': n_eval,
        'godziny_analizy': godziny_analizy,
        'godziny_opadu': godziny_opadu,
        'godziny_przegapione': godziny_przegapione,
        'godziny_puste': godziny_puste,
        'skutecznosc_oslony': skutecznosc_oslony,
        'trafnosc_alarmu': trafnosc_alarmu,
        'dokladnosc_poziomow': dokladnosc_poziomow,
        'tp': tp, 'fn': fn, 'tn': tn,
        'wczesnie': wczesnie, 'pozno': pozno, 'falszywe': falszywe,
        'poziom_total': poziom_total,
        'poziom_trafienia': poziom_trafienia,
        'horyzont_trafienia': horyzont_trafienia,
    }


# ==========================================================================
# PASEK POSTĘPU W TERMINALU
# ==========================================================================
def _pasek_postepu(ulamek, dlugosc=32):
    ulamek = max(0.0, min(1.0, ulamek))
    wypelnione = int(round(ulamek * dlugosc))
    return '[' + '#' * wypelnione + '-' * (dlugosc - wypelnione) + f'] {ulamek * 100:5.1f}%'


# ==========================================================================
# EXCEL
# ==========================================================================
def zapisz_excel(wyniki, globalne):
    os.makedirs(FOLDER_WYNIKOW, exist_ok=True)
    wb = Workbook()

    # --- Zakładka "Wyniki_lokalizacje" ---
    ws = wb.active
    ws.title = 'Wyniki_lokalizacje'
    naglowki = [
        'Lokalizacja', 'Rok', 'Krok próbkowania (min)', 'Próbki ocenione', 'Dni analizy',
        'Godziny opadu zimowego', 'Skuteczność osłony (%)', 'Trafność alarmu (%)',
        'Dokładność poziomu 0-3 (%)', 'Godziny przegapione', 'Godziny pustego grzania',
        'Za wcześnie (próbki)', 'Za późno (próbki)', 'Czysto fałszywe (próbki)',
    ] + [f'Dokł. horyzont krok {h + 1} (%)' for h in range(HORYZONT)] + [
        f'Dokł. poziom {p} (%)' for p in range(4)
    ]
    ustaw_naglowek(ws, 1, naglowki)

    for i, w in enumerate(wyniki, start=2):
        poziom_dokladnosc = [
            (w['poziom_trafienia'][p] / w['poziom_total'][p] * 100) if w['poziom_total'][p] > 0 else None
            for p in range(4)
        ]
        horyzont_dokladnosc = [(w['horyzont_trafienia'][h] / w['n_eval'] * 100) for h in range(HORYZONT)]
        wartosci = [
            w['lokalizacja'], w['rok'], round(w['krok_h'] * 60, 1), w['n_eval'],
            round(w['godziny_analizy'] / 24, 1), round(w['godziny_opadu'], 2),
            round(w['skutecznosc_oslony'], 1) if pd.notna(w['skutecznosc_oslony']) else None,
            round(w['trafnosc_alarmu'], 1) if pd.notna(w['trafnosc_alarmu']) else None,
            round(w['dokladnosc_poziomow'], 1),
            round(w['godziny_przegapione'], 2), round(w['godziny_puste'], 2),
            w['wczesnie'], w['pozno'], w['falszywe'],
        ] + [round(v, 1) for v in horyzont_dokladnosc] + [
            round(v, 1) if v is not None else None for v in poziom_dokladnosc
        ]
        for j, wartosc in enumerate(wartosci, start=1):
            komorka = ws.cell(row=i, column=j, value=wartosc)
            komorka.font = FONT_ZWYKLY
            komorka.border = OBRAMOWANIE_CIENKIE
            if j in (1,):
                komorka.alignment = WYROWNANIE_SRODEK

    ostatni_wiersz = len(wyniki) + 1
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{ws.cell(row=1, column=len(naglowki)).coordinate[:-1]}{ostatni_wiersz}'

    for kolumna_nazwa, litera in (('Skuteczność osłony (%)', 'G'), ('Trafność alarmu (%)', 'H'),
                                    ('Dokładność poziomu 0-3 (%)', 'I')):
        skala = ColorScaleRule(start_type='min', start_color='F8696B',
                                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                                end_type='max', end_color='63BE7B')
        ws.conditional_formatting.add(f'{litera}2:{litera}{ostatni_wiersz}', skala)

    autoszerokosc(ws)

    # --- Zakładka "Podsumowanie_ogolne" ---
    ws2 = wb.create_sheet('Podsumowanie_ogolne')
    ws2.cell(row=1, column=1, value='Skuteczność przewidywanie_opadow.py na wszystkich próbkach pogodowych').font = \
        Font(name=FONT_NAZWA, bold=True, size=14)

    n_lokalizacji = len(wyniki)
    n_eval_total = globalne['n_eval']
    godziny_opadu_total = sum(w['godziny_opadu'] for w in wyniki)
    tp_total = globalne['tp']
    fn_total = globalne['fn']
    wczesnie_total = globalne['wczesnie']
    pozno_total = globalne['pozno']
    falszywe_total = globalne['falszywe']

    # Globalne (pooled) - ważone rzeczywistą liczbą godzin z każdego pliku,
    # bo krok próbkowania różni się między plikami (15 vs 60 min).
    godziny_wykryte_total = sum(w['tp'] * w['krok_h'] for w in wyniki)
    godziny_przegapione_total = sum(w['fn'] * w['krok_h'] for w in wyniki)
    godziny_puste_total = sum((w['wczesnie'] + w['pozno'] + w['falszywe']) * w['krok_h'] for w in wyniki)
    skutecznosc_globalna = (godziny_wykryte_total / godziny_opadu_total * 100) if godziny_opadu_total > 0 else 0.0
    wszystkie_alarmy_total = tp_total + wczesnie_total + pozno_total + falszywe_total
    trafnosc_globalna = (tp_total / wszystkie_alarmy_total * 100) if wszystkie_alarmy_total > 0 else 0.0
    poziom_total_suma = sum((w['poziom_trafienia'].sum() for w in wyniki))
    dokladnosc_poziomow_globalna = poziom_total_suma / n_eval_total * 100

    # Macro-average (prosta średnia po lokalizacjach - każda lokalizacja liczy
    # się tyle samo, niezależnie od tego, ile miała godzin opadu) - pokazuje
    # SPÓJNOŚĆ skuteczności między lokalizacjami, w odróżnieniu od globalnej
    # (pooled) wartości wyżej, którą dominują lokalizacje z długimi opadami.
    skutecznosci = [w['skutecznosc_oslony'] for w in wyniki if pd.notna(w['skutecznosc_oslony'])]
    trafnosci = [w['trafnosc_alarmu'] for w in wyniki if pd.notna(w['trafnosc_alarmu'])]
    macro_skutecznosc = float(np.mean(skutecznosci)) if skutecznosci else float('nan')
    macro_trafnosc = float(np.mean(trafnosci)) if trafnosci else float('nan')

    najlepsza = max(wyniki, key=lambda w: (w['skutecznosc_oslony'] if pd.notna(w['skutecznosc_oslony']) else -1))
    najgorsza = min(wyniki, key=lambda w: (w['skutecznosc_oslony'] if pd.notna(w['skutecznosc_oslony']) else 101))

    wiersze_tekst = [
        ('', ''),
        ('Liczba ocenionych plików pogodowych (lokalizacja x rok)', n_lokalizacji),
        ('Łączna liczba ocenionych próbek', n_eval_total),
        ('Łączna liczba godzin opadu zimowego w danych', round(godziny_opadu_total, 1)),
        ('', ''),
        ('SKUTECZNOŚĆ OSŁONY (globalna, ważona godzinami opadu)', f'{skutecznosc_globalna:.1f}%'),
        ('SKUTECZNOŚĆ OSŁONY (średnia po lokalizacjach, macro-avg)', f'{macro_skutecznosc:.1f}%'),
        ('TRAFNOŚĆ ALARMU (globalna, ważona liczbą alarmów)', f'{trafnosc_globalna:.1f}%'),
        ('TRAFNOŚĆ ALARMU (średnia po lokalizacjach, macro-avg)', f'{macro_trafnosc:.1f}%'),
        ('DOKŁADNOŚĆ dokładnego poziomu 0-3 (globalna)', f'{dokladnosc_poziomow_globalna:.1f}%'),
        ('', ''),
        ('Godziny opadu wykryte (globalnie)', round(godziny_wykryte_total, 1)),
        ('Godziny opadu przegapione (globalnie)', round(godziny_przegapione_total, 1)),
        ('Godziny pustego grzania - fałszywy alarm (globalnie)', round(godziny_puste_total, 1)),
        ('', ''),
        (f'Najlepsza lokalizacja (skuteczność osłony)',
         f"{najlepsza['lokalizacja']} {najlepsza['rok']}: {najlepsza['skutecznosc_oslony']:.1f}%"),
        (f'Najgorsza lokalizacja (skuteczność osłony)',
         f"{najgorsza['lokalizacja']} {najgorsza['rok']}: {najgorsza['skutecznosc_oslony']:.1f}%"),
    ]
    for i, (etykieta, wartosc) in enumerate(wiersze_tekst, start=3):
        ws2.cell(row=i, column=1, value=etykieta).font = FONT_POGRUBIONY if etykieta else FONT_ZWYKLY
        ws2.cell(row=i, column=2, value=wartosc).font = FONT_ZWYKLY

    # Tabela pomocnicza pod wykres: dokładność wg horyzontu (globalna, pooled)
    wiersz_tabeli_horyzont = len(wiersze_tekst) + 6
    ws2.cell(row=wiersz_tabeli_horyzont - 1, column=1, value='Dokładność wg horyzontu prognozy (globalna, pooled)').font = FONT_POGRUBIONY
    ws2.cell(row=wiersz_tabeli_horyzont, column=1, value='Krok naprzód').font = FONT_NAGLOWEK
    ws2.cell(row=wiersz_tabeli_horyzont, column=1).fill = FILL_NAGLOWEK
    ws2.cell(row=wiersz_tabeli_horyzont, column=2, value='Dokładność (%)').font = FONT_NAGLOWEK
    ws2.cell(row=wiersz_tabeli_horyzont, column=2).fill = FILL_NAGLOWEK

    horyzont_trafienia_total = sum((w['horyzont_trafienia'] for w in wyniki))
    for h in range(HORYZONT):
        dokladnosc_h = horyzont_trafienia_total[h] / n_eval_total * 100
        ws2.cell(row=wiersz_tabeli_horyzont + 1 + h, column=1, value=h + 1).font = FONT_ZWYKLY
        ws2.cell(row=wiersz_tabeli_horyzont + 1 + h, column=2, value=round(float(dokladnosc_h), 2)).font = FONT_ZWYKLY

    chart = LineChart()
    chart.title = 'Dokładność prognozy w funkcji horyzontu (wszystkie lokalizacje, pooled)'
    chart.x_axis.title = 'Krok prognozy naprzód (jednostka = krok źródłowy pliku, 15 lub 60 min)'
    chart.y_axis.title = 'Dokładność (%)'
    chart.width = 24
    chart.height = 12
    dane_serii = Reference(ws2, min_col=2, min_row=wiersz_tabeli_horyzont,
                            max_row=wiersz_tabeli_horyzont + HORYZONT)
    kategorie = Reference(ws2, min_col=1, min_row=wiersz_tabeli_horyzont + 1,
                           max_row=wiersz_tabeli_horyzont + HORYZONT)
    chart.add_data(dane_serii, titles_from_data=True)
    chart.set_categories(kategorie)
    ws2.add_chart(chart, f'D{wiersz_tabeli_horyzont - 1}')

    autoszerokosc(ws2)

    # --- Zakładka "Definicje_poziomow" ---
    ws3 = wb.create_sheet('Definicje_poziomow')
    ustaw_naglowek(ws3, 1, ['Poziom', 'Nazwa', 'Próg (definicja ground truth)'])
    for i, (lvl, (nazwa, prog)) in enumerate(INTENSITY_DEFS.items(), start=2):
        for j, wartosc in enumerate([lvl, nazwa, prog], start=1):
            komorka = ws3.cell(row=i, column=j, value=wartosc)
            komorka.font = FONT_ZWYKLY
            komorka.border = OBRAMOWANIE_CIENKIE
    autoszerokosc(ws3)

    wb.save(SCIEZKA_XLSX)


# ==========================================================================
# GŁÓWNY PRZEBIEG
# ==========================================================================
def main():
    pliki = sorted(glob.glob(os.path.join(FOLDER_POGODA, '*.csv')))
    n_plikow = len(pliki)
    if n_plikow == 0:
        print(f"Brak plików CSV w {FOLDER_POGODA}")
        return

    print(f"Znaleziono {n_plikow} plików pogodowych do oceny.\n")

    wyniki = []
    globalne = dict(tp=0, fn=0, tn=0, wczesnie=0, pozno=0, falszywe=0, n_eval=0)

    czas_startu = time.time()
    for idx, plik in enumerate(pliki, start=1):
        lokalizacja, krok_plik_min, rok = _nazwa_lokalizacji(plik)
        etykieta = f"{lokalizacja} {rok}" if rok else lokalizacja

        def callback(zrobione, razem, idx=idx, etykieta=etykieta):
            ulamek_pliku = zrobione / razem if razem else 1.0
            ulamek_calosci = ((idx - 1) + ulamek_pliku) / n_plikow
            pasek = _pasek_postepu(ulamek_calosci)
            sys.stdout.write(f"\r{pasek}  plik {idx}/{n_plikow}: {etykieta:<32}")
            sys.stdout.flush()

        wynik = oceniaj_plik(plik, callback_postepu=callback)
        wynik['lokalizacja'] = lokalizacja
        wynik['rok'] = rok
        wynik['plik'] = os.path.basename(plik)
        wyniki.append(wynik)

        for klucz in ('tp', 'fn', 'tn', 'wczesnie', 'pozno', 'falszywe', 'n_eval'):
            globalne[klucz] += wynik[klucz]

        sys.stdout.write('\n')
        skut = f"{wynik['skutecznosc_oslony']:.1f}%" if pd.notna(wynik['skutecznosc_oslony']) else 'n/d (brak opadu w pliku)'
        traf = f"{wynik['trafnosc_alarmu']:.1f}%" if pd.notna(wynik['trafnosc_alarmu']) else 'n/d'
        print(f"   -> osłona: {skut:<10}  trafność alarmu: {traf:<10}  "
              f"dokł. poziomów: {wynik['dokladnosc_poziomow']:.1f}%  "
              f"(krok {wynik['krok_h'] * 60:.0f} min, {wynik['n_eval']} próbek)")

    czas_calkowity = time.time() - czas_startu
    print(f"\n{_pasek_postepu(1.0)}  Zakończono ocenę {n_plikow} plików w {czas_calkowity:.1f}s.\n")

    zapisz_excel(wyniki, globalne)

    godziny_opadu_total = sum(w['godziny_opadu'] for w in wyniki)
    godziny_wykryte_total = sum(w['tp'] * w['krok_h'] for w in wyniki)
    skutecznosc_globalna = (godziny_wykryte_total / godziny_opadu_total * 100) if godziny_opadu_total > 0 else 0.0
    wszystkie_alarmy_total = globalne['tp'] + globalne['wczesnie'] + globalne['pozno'] + globalne['falszywe']
    trafnosc_globalna = (globalne['tp'] / wszystkie_alarmy_total * 100) if wszystkie_alarmy_total > 0 else 0.0

    print("=" * 75)
    print("PODSUMOWANIE ZBIORCZE (wszystkie lokalizacje)")
    print("=" * 75)
    print(f"Skuteczność osłony (globalna, ważona godzinami opadu): {skutecznosc_globalna:.1f}%")
    print(f"Trafność alarmu (globalna, ważona liczbą alarmów):     {trafnosc_globalna:.1f}%")
    print(f"Łączna liczba godzin opadu zimowego w danych:           {godziny_opadu_total:.1f} h")
    print("=" * 75)
    print(f"\nSzczegóły zapisane w: {SCIEZKA_XLSX}")


if __name__ == '__main__':
    main()
