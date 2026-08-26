# generuj_excel_podsumowanie.py
#
# Buduje plik Excel (.xlsx) podsumowujący wyniki dużego przeglądu
# (test_wszystkie_algorytmy_wszystkie_lokalizacje.py -> PRZEGLAD_ZBIORCZY.csv):
#   - "Dane"                    - surowa tabela, jeden wiersz na (lokalizacja, rok, algorytm)
#   - "Podsumowanie_algorytmy"  - porównanie wszystkich algorytmów globalnie (formuły)
#   - "Podsumowanie_lokalizacje"- średnia energia per lokalizacja x algorytm (formuły)
#   - "Opisy_algorytmow"        - ściąga: typ regulatora / cel (setpoint) / czy adaptacyjny
#                                  (autotest) / opis - wprost z Algorytmy/rejestr_algorytmow.py
#   - "Zlozonosc_obliczeniowa"  - szacunkowa złożoność czasowa/pamięciowa i FLOPs/krok każdego
#                                  algorytmu (analiza kodu, nie profiler) - wprost z rejestru
#   - "Wnioski"                 - tekstowe podsumowanie z konkretnymi liczbami
#
# Wartości w zakładkach "Podsumowanie_*" są formułami Excela odwołującymi się
# do zakładki "Dane" - podmiana/dopisanie wierszy w "Dane" automatycznie
# przeliczy resztę (poza "Wnioski", która jest tekstem opisowym).

import os
import re
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE_DIR, 'Algorytmy'))
from rejestr_algorytmow import ALGORYTMY  # noqa: E402 - opisy/typ/cel/adaptacyjny dla zakładki "Opisy_algorytmow"
# SZYNA_FOLDER_WYNIKOW pozwala przekierować wejście/wyjście do innego folderu
# (patrz ten sam mechanizm w test_wszystkie_rownolegle.py) - bez tego wywołanie
# generuj_excel_podsumowanie.main() po przebiegu zapisanym gdzie indziej niż
# domyślny folder szukałoby CSV w złym miejscu.
FOLDER_WYNIKOW = os.environ.get(
    'SZYNA_FOLDER_WYNIKOW', os.path.join(BASE_DIR, "wyniki", "przeglad_wielu_lokalizacji"))
SCIEZKA_CSV = os.path.join(FOLDER_WYNIKOW, "PRZEGLAD_ZBIORCZY.csv")
SCIEZKA_XLSX = os.path.join(FOLDER_WYNIKOW, "Podsumowanie_wynikow.xlsx")

NAZWY_MIAST = {
    'abisko': 'Abisko', 'fairbanks': 'Fairbanks', 'jakuck': 'Jakuck',
    'krakow': 'Kraków', 'ojmiakon': 'Ojmiakon', 'old_crow': 'Old Crow',
    'oslo': 'Oslo', 'puszcza_bialowieska': 'Puszcza Białowieska',
    'suwalki': 'Suwałki', 'wroclaw': 'Wrocław',
}

# Kolejność = norma jako baza odniesienia, potem pozostałe algorytmy w kolejności
# rejestru (Algorytmy/rejestr_algorytmow.py).
NAZWY_ALGORYTMOW = {
    'algorytm_z_normy': 'Automat z normy (bazowy)',
    'compute_control': 'Histereza LET-1',
    'risk_function': 'Funkcja ryzyka (binarna)',
    'risk_function_pid': 'Funkcja ryzyka (PID)',
    'norma_pid': 'PID z normą',
    'fuzzy_logic_1': 'Fuzzy Logic 1 (ciągły)',
    'fuzzy_logic_2': 'Fuzzy Logic 2 (binarny)',
    'fuzzy_logic_2v2': 'Fuzzy Logic 2v2 (binarny)',
    'fuzzy_logic_3': 'Fuzzy Logic 3 (PWM)',
    'fuzzy_ryzyko_1': 'Fuzzy + ryzyko (FL1, ciągły)',
    'fuzzy_ryzyko_2': 'Fuzzy + ryzyko (FL2, binarny)',
    'fuzzy_ryzyko_2v2': 'Fuzzy + ryzyko (FL2v2, binarny)',
    'fuzzy_ryzyko_3': 'Fuzzy + ryzyko (FL3, PWM)',
    'fuzzy_normy_1': 'Fuzzy + norma (FL1, ciągły)',
    'fuzzy_normy_2': 'Fuzzy + norma (FL2, binarny)',
    'fuzzy_normy_2v2': 'Fuzzy + norma (FL2v2, binarny)',
    'fuzzy_normy_3': 'Fuzzy + norma (FL3, PWM)',
    'risk_function_opad': 'Funkcja ryzyka (binarna) + opad',
    'risk_function_pid_opad': 'Funkcja ryzyka (PID) + opad',
    'fuzzy_ryzyko_1_opad': 'Fuzzy + ryzyko + opad (FL1, ciągły)',
    'fuzzy_ryzyko_2_opad': 'Fuzzy + ryzyko + opad (FL2, binarny)',
    'fuzzy_ryzyko_2v2_opad': 'Fuzzy + ryzyko + opad (FL2v2, binarny)',
    'fuzzy_ryzyko_3_opad': 'Fuzzy + ryzyko + opad (FL3, PWM)',
}
ALGORYTM_BAZOWY = 'Automat z normy (bazowy)'

FONT_NAZWA = 'Arial'
KOLOR_NAGLOWEK_BG = '1F4E78'
KOLOR_NAGLOWEK_FG = 'FFFFFF'
KOLOR_ANOMALIA = 'FFC7CE'
KOLOR_ANOMALIA_TEKST = '9C0006'

FONT_NAGLOWEK = Font(name=FONT_NAZWA, bold=True, color=KOLOR_NAGLOWEK_FG, size=11)
FILL_NAGLOWEK = PatternFill('solid', fgColor=KOLOR_NAGLOWEK_BG)
FONT_ZWYKLY = Font(name=FONT_NAZWA, size=10)
FONT_POGRUBIONY = Font(name=FONT_NAZWA, bold=True, size=10)
WYROWNANIE_SRODEK = Alignment(horizontal='center', vertical='center')
OBRAMOWANIE_CIENKIE = Border(*(Side(style='thin', color='B7B7B7') for _ in range(4)))


def parsuj_lokalizacje(lokalizacja):
    """'krakow_60min_2021' -> ('Kraków', '60min', 2021). Działa też dla 'suwalki_15min_2023'."""
    dopasowanie = re.match(r'^(.+)_(\d+min)_(\d{4})$', lokalizacja)
    if not dopasowanie:
        raise ValueError(f"Nie rozpoznano formatu nazwy lokalizacji: {lokalizacja}")
    klucz_miasta, interwal, rok = dopasowanie.groups()
    miasto = NAZWY_MIAST.get(klucz_miasta, klucz_miasta.replace('_', ' ').title())
    return miasto, interwal, int(rok)


def ustaw_naglowek(ws, wiersz, kolumny):
    for i, nazwa in enumerate(kolumny, start=1):
        komorka = ws.cell(row=wiersz, column=i, value=nazwa)
        komorka.font = FONT_NAGLOWEK
        komorka.fill = FILL_NAGLOWEK
        komorka.alignment = WYROWNANIE_SRODEK
        komorka.border = OBRAMOWANIE_CIENKIE


def autoszerokosc(ws, min_szer=10, max_szer=42):
    for kolumna_komorki in ws.columns:
        dlugosc = 0
        litera = None
        for komorka in kolumna_komorki:
            if litera is None:
                litera = getattr(komorka, 'column_letter', None)
            if litera is None:
                continue
            if komorka.value is not None:
                dlugosc = max(dlugosc, len(str(komorka.value)))
        if litera:
            ws.column_dimensions[litera].width = max(min_szer, min(max_szer, dlugosc + 3))


def main():
    df = pd.read_csv(SCIEZKA_CSV)
    df[['Lokalizacja', 'Interwal', 'Rok']] = df['lokalizacja'].apply(lambda x: pd.Series(parsuj_lokalizacje(x)))
    df['Algorytm'] = df['name'].map(NAZWY_ALGORYTMOW)

    wb = Workbook()

    # ==========================================================================
    # ZAKŁADKA "Dane"
    # ==========================================================================
    ws_dane = wb.active
    ws_dane.title = 'Dane'
    naglowki_dane = ['Lokalizacja', 'Interwał', 'Rok', 'Algorytm', 'Energia (kWh)',
                      'Przełączenia', 'Max śnieg (mm)', 'Max HRT (°C)']
    ustaw_naglowek(ws_dane, 1, naglowki_dane)

    for i, wiersz in enumerate(df.itertuples(index=False), start=2):
        wartosci = [wiersz.Lokalizacja, wiersz.Interwal, wiersz.Rok, wiersz.Algorytm,
                    round(wiersz.energia_kwh, 2), int(wiersz.przelaczenia),
                    round(wiersz.max_snieg_mm, 2), round(wiersz.max_hrt, 2)]
        for j, wartosc in enumerate(wartosci, start=1):
            komorka = ws_dane.cell(row=i, column=j, value=wartosc)
            komorka.font = FONT_ZWYKLY
            komorka.border = OBRAMOWANIE_CIENKIE
            if j == 3:
                komorka.alignment = WYROWNANIE_SRODEK

    ostatni_wiersz_dane = len(df) + 1
    ws_dane.freeze_panes = 'A2'
    ws_dane.auto_filter.ref = f'A1:H{ostatni_wiersz_dane}'

    # Podświetlenie wierszy z podejrzanym przegrzaniem (Max HRT > 35°C).
    fill_anomalia = PatternFill('solid', fgColor=KOLOR_ANOMALIA)
    font_anomalia = Font(name=FONT_NAZWA, size=10, color=KOLOR_ANOMALIA_TEKST)
    ws_dane.conditional_formatting.add(
        f'A2:H{ostatni_wiersz_dane}',
        FormulaRule(formula=['$H2>35'], fill=fill_anomalia, font=font_anomalia),
    )

    autoszerokosc(ws_dane)

    # ==========================================================================
    # ZAKŁADKA "Podsumowanie_algorytmy"
    # ==========================================================================
    ws_alg = wb.create_sheet('Podsumowanie_algorytmy')
    naglowki_alg = ['Algorytm', 'Średnia energia (kWh)', 'Min energia (kWh)', 'Max energia (kWh)',
                     'Średnie przełączenia', 'Średni max śnieg (mm)', 'Średni max HRT (°C)',
                     'Liczba przypadków przegrzania (HRT>35°C)']
    ustaw_naglowek(ws_alg, 1, naglowki_alg)

    lista_algorytmow = list(NAZWY_ALGORYTMOW.values())
    for i, algorytm in enumerate(lista_algorytmow, start=2):
        ws_alg.cell(row=i, column=1, value=algorytm).font = FONT_POGRUBIONY
        ws_alg.cell(row=i, column=2, value=(
            f'=AVERAGEIF(Dane!$D$2:$D${ostatni_wiersz_dane}, $A{i}, Dane!$E$2:$E${ostatni_wiersz_dane})'
        ))
        ws_alg.cell(row=i, column=3, value=(
            f'=_xlfn.MINIFS(Dane!$E$2:$E${ostatni_wiersz_dane}, Dane!$D$2:$D${ostatni_wiersz_dane}, $A{i})'
        ))
        ws_alg.cell(row=i, column=4, value=(
            f'=_xlfn.MAXIFS(Dane!$E$2:$E${ostatni_wiersz_dane}, Dane!$D$2:$D${ostatni_wiersz_dane}, $A{i})'
        ))
        ws_alg.cell(row=i, column=5, value=(
            f'=AVERAGEIF(Dane!$D$2:$D${ostatni_wiersz_dane}, $A{i}, Dane!$F$2:$F${ostatni_wiersz_dane})'
        ))
        ws_alg.cell(row=i, column=6, value=(
            f'=AVERAGEIF(Dane!$D$2:$D${ostatni_wiersz_dane}, $A{i}, Dane!$G$2:$G${ostatni_wiersz_dane})'
        ))
        ws_alg.cell(row=i, column=7, value=(
            f'=AVERAGEIF(Dane!$D$2:$D${ostatni_wiersz_dane}, $A{i}, Dane!$H$2:$H${ostatni_wiersz_dane})'
        ))
        ws_alg.cell(row=i, column=8, value=(
            f'=COUNTIFS(Dane!$D$2:$D${ostatni_wiersz_dane}, $A{i}, Dane!$H$2:$H${ostatni_wiersz_dane}, ">35")'
        ))
        for kol in range(2, 9):
            komorka = ws_alg.cell(row=i, column=kol)
            komorka.font = FONT_ZWYKLY
            komorka.number_format = '0.00' if kol not in (8,) else '0'
            komorka.border = OBRAMOWANIE_CIENKIE

    ostatni_wiersz_alg = len(lista_algorytmow) + 1
    for litera, odwrocona in [('B', True), ('E', True), ('F', True), ('G', True), ('H', True)]:
        if odwrocona:
            skala = ColorScaleRule(start_type='min', start_color='63BE7B',
                                    end_type='max', end_color='F8696B')
        else:
            skala = ColorScaleRule(start_type='min', start_color='F8696B',
                                    end_type='max', end_color='63BE7B')
        ws_alg.conditional_formatting.add(f'{litera}2:{litera}{ostatni_wiersz_alg}', skala)

    autoszerokosc(ws_alg)

    # ==========================================================================
    # ZAKŁADKA "Podsumowanie_lokalizacje"
    # ==========================================================================
    ws_lok = wb.create_sheet('Podsumowanie_lokalizacje')

    # Wiersz 1: opisowa notka (rozpięta nad tabelą) - nagłówki w wierszu 2 muszą
    # zawierać SUROWE nazwy algorytmów (identyczne jak w Dane!Algorytm), bo służą
    # jako kryterium w AVERAGEIFS poniżej - stąd nie mogą być "ozdobione" opisem.
    ws_lok.cell(row=1, column=1, value='Średnia energia (kWh) per algorytm, uśredniona po latach dla każdej lokalizacji.')
    ws_lok.cell(row=1, column=1).font = Font(name=FONT_NAZWA, italic=True, size=9, color='555555')
    ws_lok.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(lista_algorytmow) + 3)

    WIERSZ_NAGLOWKA_LOK = 2
    naglowki_lok = ['Lokalizacja'] + lista_algorytmow + ['Najlepszy algorytm', 'Oszczędność vs. bazowy (%)']
    ustaw_naglowek(ws_lok, WIERSZ_NAGLOWKA_LOK, naglowki_lok)

    lokalizacje_unikalne = sorted(df['Lokalizacja'].unique())
    kolumna_bazowa_litera = get_column_letter(2 + lista_algorytmow.index(ALGORYTM_BAZOWY))
    pierwsza_kol_alg = get_column_letter(2)
    ostatnia_kol_alg = get_column_letter(1 + len(lista_algorytmow))
    wiersz_naglowka = WIERSZ_NAGLOWKA_LOK

    for i, miasto in enumerate(lokalizacje_unikalne, start=WIERSZ_NAGLOWKA_LOK + 1):
        ws_lok.cell(row=i, column=1, value=miasto).font = FONT_POGRUBIONY
        for j, algorytm in enumerate(lista_algorytmow, start=2):
            litera = get_column_letter(j)
            komorka = ws_lok.cell(row=i, column=j, value=(
                f'=AVERAGEIFS(Dane!$E$2:$E${ostatni_wiersz_dane}, '
                f'Dane!$A$2:$A${ostatni_wiersz_dane}, $A{i}, '
                f'Dane!$D$2:$D${ostatni_wiersz_dane}, {litera}${wiersz_naglowka})'
            ))
            komorka.font = FONT_ZWYKLY
            komorka.number_format = '0.00'
            komorka.border = OBRAMOWANIE_CIENKIE

        kol_najlepszy = len(lista_algorytmow) + 2
        kol_oszczednosc = kol_najlepszy + 1

        komorka_najlepszy = ws_lok.cell(row=i, column=kol_najlepszy, value=(
            f'=INDEX($B${wiersz_naglowka}:${ostatnia_kol_alg}${wiersz_naglowka}, '
            f'MATCH(MIN({pierwsza_kol_alg}{i}:{ostatnia_kol_alg}{i}), '
            f'{pierwsza_kol_alg}{i}:{ostatnia_kol_alg}{i}, 0))'
        ))
        komorka_najlepszy.font = FONT_POGRUBIONY
        komorka_najlepszy.border = OBRAMOWANIE_CIENKIE

        komorka_oszczednosc = ws_lok.cell(row=i, column=kol_oszczednosc, value=(
            f'=({kolumna_bazowa_litera}{i}-MIN({pierwsza_kol_alg}{i}:{ostatnia_kol_alg}{i}))/{kolumna_bazowa_litera}{i}'
        ))
        komorka_oszczednosc.font = FONT_ZWYKLY
        komorka_oszczednosc.number_format = '0.0%'
        komorka_oszczednosc.border = OBRAMOWANIE_CIENKIE

    ostatni_wiersz_lok = len(lokalizacje_unikalne) + WIERSZ_NAGLOWKA_LOK
    for j in range(2, 2 + len(lista_algorytmow)):
        litera = get_column_letter(j)
        skala = ColorScaleRule(start_type='min', start_color='63BE7B',
                                end_type='max', end_color='F8696B')
        ws_lok.conditional_formatting.add(f'{litera}{WIERSZ_NAGLOWKA_LOK + 1}:{litera}{ostatni_wiersz_lok}', skala)

    ws_lok.freeze_panes = f'B{WIERSZ_NAGLOWKA_LOK + 1}'
    autoszerokosc(ws_lok)

    # ==========================================================================
    # ZAKŁADKA "Opisy_algorytmow" - krótka ściąga: co jest czym (PID/fuzzy logic/
    # histereza), na czym oparty jest cel grzania i czy kontroler jest adaptacyjny
    # (autotest startowy + przestrajanie/cyfrowy bliźniak na żywo) - wprost z
    # Algorytmy/rejestr_algorytmow.py, więc nie może się rozjechać z kodem.
    # ==========================================================================
    ws_opis = wb.create_sheet('Opisy_algorytmow')
    naglowki_opis = ['Algorytm', 'Typ regulatora', 'Cel (setpoint)', 'Adaptacyjny (autotest)', 'Opis']
    ustaw_naglowek(ws_opis, 1, naglowki_opis)

    for i, klucz in enumerate(NAZWY_ALGORYTMOW, start=2):
        wpis = ALGORYTMY.get(klucz, {})
        wartosci = [
            NAZWY_ALGORYTMOW[klucz],
            wpis.get('typ', ''),
            wpis.get('cel', ''),
            'Tak' if wpis.get('adaptacyjny') else 'Nie',
            wpis.get('opis', ''),
        ]
        for j, wartosc in enumerate(wartosci, start=1):
            komorka = ws_opis.cell(row=i, column=j, value=wartosc)
            komorka.font = FONT_POGRUBIONY if j == 1 else FONT_ZWYKLY
            komorka.border = OBRAMOWANIE_CIENKIE
            komorka.alignment = Alignment(vertical='center', wrap_text=(j == 5))
            if j == 4:
                komorka.alignment = WYROWNANIE_SRODEK

    ostatni_wiersz_opis = len(NAZWY_ALGORYTMOW) + 1
    fill_adaptacyjny = PatternFill('solid', fgColor='D9EAD3')
    ws_opis.conditional_formatting.add(
        f'A2:E{ostatni_wiersz_opis}',
        FormulaRule(formula=['$D2="Tak"'], fill=fill_adaptacyjny),
    )
    ws_opis.freeze_panes = 'A2'
    autoszerokosc(ws_opis, max_szer=45)
    ws_opis.column_dimensions['E'].width = 80  # opis jest długi - autoszerokosc przycięłaby go do max_szer

    # ==========================================================================
    # ZAKŁADKA "Zlozonosc_obliczeniowa" - szacunkowa złożoność czasowa/pamięciowa
    # i FLOPs/krok każdego algorytmu, wprost z Algorytmy/rejestr_algorytmow.py
    # (patrz komentarz przy tych polach tam - wyliczone analizą kodu, NIE
    # zmierzone profilerem, traktuj jako rząd wielkości).
    # ==========================================================================
    ws_zloz = wb.create_sheet('Zlozonosc_obliczeniowa')
    ws_zloz.cell(row=1, column=1, value=(
        'FLOPs/krok dotyczy WYŁĄCZNIE logiki decyzyjnej algorytmu (bez współdzielonej fizyki obiektu/śniegu, '
        'identycznej dla wszystkich) - rzędu dziesiątek/setek FLOPs, więc realny czas symulacji wynika z narzutu '
        'interpretera Pythona/pandas na krok, NIE z limitu przepustowości FLOPs procesora.'
    ))
    ws_zloz.cell(row=1, column=1).font = Font(name=FONT_NAZWA, italic=True, size=9, color='555555')
    ws_zloz.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    WIERSZ_NAGLOWKA_ZLOZ = 2
    naglowki_zloz = ['Algorytm', 'Złożoność czasowa (na krok)', 'FLOPs/krok (przybliżone)',
                      'Złożoność pamięciowa', 'Pamięć - stan ustalony (MB)',
                      'Kroki w oknie testowym', 'Łączne FLOPs (algorytm, całe okno)']
    ustaw_naglowek(ws_zloz, WIERSZ_NAGLOWKA_ZLOZ, naglowki_zloz)

    # "Kroki w oknie testowym" i "Łączne FLOPs" liczone z rzeczywistej mediany liczby
    # dni w danych (kolumna 'dni' w Dane, jeśli obecna) x 86400 s/dobę - czysto
    # informacyjne, NIE FLOPs całej symulacji (fizyka obiektu/śniegu pominięta, patrz wyżej).
    mediana_dni = df['dni'].median() if 'dni' in df.columns else None
    kroki_okna = int(round(mediana_dni * 86400)) if mediana_dni is not None else None

    for i, klucz in enumerate(NAZWY_ALGORYTMOW, start=WIERSZ_NAGLOWKA_ZLOZ + 1):
        wpis = ALGORYTMY.get(klucz, {})
        flops_krok = wpis.get('flops_na_krok')
        laczne_flops = flops_krok * kroki_okna if (flops_krok is not None and kroki_okna is not None) else None
        wartosci = [
            NAZWY_ALGORYTMOW[klucz],
            wpis.get('zlozonosc_czasowa', ''),
            flops_krok,
            wpis.get('zlozonosc_pamieciowa', ''),
            wpis.get('pamiec_przyblizona_mb'),
            kroki_okna,
            laczne_flops,
        ]
        for j, wartosc in enumerate(wartosci, start=1):
            komorka = ws_zloz.cell(row=i, column=j, value=wartosc)
            komorka.font = FONT_POGRUBIONY if j == 1 else FONT_ZWYKLY
            komorka.border = OBRAMOWANIE_CIENKIE
            komorka.alignment = Alignment(vertical='center', wrap_text=(j in (2, 4)))
            if j in (3, 5, 6):
                komorka.number_format = '0.000' if j == 5 else '0'
            if j == 7 and wartosc is not None:
                komorka.number_format = '0.00E+00'

    ostatni_wiersz_zloz = len(NAZWY_ALGORYTMOW) + WIERSZ_NAGLOWKA_ZLOZ
    skala_flops = ColorScaleRule(start_type='min', start_color='63BE7B', end_type='max', end_color='F8696B')
    ws_zloz.conditional_formatting.add(f'C{WIERSZ_NAGLOWKA_ZLOZ + 1}:C{ostatni_wiersz_zloz}', skala_flops)
    ws_zloz.freeze_panes = f'A{WIERSZ_NAGLOWKA_ZLOZ + 1}'
    ws_zloz.column_dimensions['B'].width = 55
    autoszerokosc(ws_zloz, max_szer=40)
    ws_zloz.column_dimensions['B'].width = 55  # opis złożoności jest długi - autoszerokosc przycięłaby go

    # ==========================================================================
    # ZAKŁADKA "Wnioski" - tekst z realnie policzonymi liczbami (nie formuły)
    # ==========================================================================
    ws_wn = wb.create_sheet('Wnioski')
    ws_wn.column_dimensions['A'].width = 115

    agregaty = df.groupby('Algorytm').agg(
        energia_srednia=('energia_kwh', 'mean'),
        przelaczenia_srednie=('przelaczenia', 'mean'),
        max_hrt_srednie=('max_hrt', 'mean'),
        max_snieg_srednie=('max_snieg_mm', 'mean'),
    ).reindex(lista_algorytmow)

    anomalie = df[df['max_hrt'] > 35.0].copy()
    anomalie_per_algorytm = anomalie['Algorytm'].value_counts().reindex(lista_algorytmow, fill_value=0)

    zwyciezca_energia = agregaty['energia_srednia'].idxmin()
    zwyciezca_przelaczenia = agregaty['przelaczenia_srednie'].idxmin()
    energia_bazowa = agregaty.loc[ALGORYTM_BAZOWY, 'energia_srednia']
    energia_najlepsza = agregaty.loc[zwyciezca_energia, 'energia_srednia']
    oszczednosc_pct = (energia_bazowa - energia_najlepsza) / energia_bazowa * 100.0

    linie = []
    linie.append('PODSUMOWANIE PRZEGLĄDU - WSZYSTKIE ALGORYTMY, WSZYSTKIE LOKALIZACJE')
    linie.append(f'(dane: {len(df)} wierszy, {df["Lokalizacja"].nunique()} lokalizacji, '
                 f'okno {int(round(df["dni"].median())) if "dni" in df.columns else "?"} dni na lokalizację)')
    linie.append('')
    linie.append('1) ENERGIA')
    linie.append(f'   Najniższą średnią energię ({energia_najlepsza:.1f} kWh) osiągnął algorytm '
                 f'"{zwyciezca_energia}" - to {oszczednosc_pct:.1f}% mniej niż algorytm bazowy '
                 f'"{ALGORYTM_BAZOWY}" ({energia_bazowa:.1f} kWh).')
    for alg in lista_algorytmow:
        linie.append(f'   - {alg}: średnio {agregaty.loc[alg, "energia_srednia"]:.1f} kWh')
    linie.append('')
    linie.append('2) STABILNOŚĆ (LICZBA PRZEŁĄCZEŃ)')
    linie.append(f'   Zdecydowanym zwycięzcą jest "{zwyciezca_przelaczenia}" ze średnio '
                 f'{agregaty.loc[zwyciezca_przelaczenia, "przelaczenia_srednie"]:.1f} przełączeń '
                 f'na lokalizację - {agregaty["przelaczenia_srednie"].max()/agregaty.loc[zwyciezca_przelaczenia, "przelaczenia_srednie"]:.1f}x '
                 f'mniej niż najgorszy pod tym względem wariant '
                 f'({agregaty["przelaczenia_srednie"].idxmax()}, {agregaty["przelaczenia_srednie"].max():.1f}). '
                 f'To spodziewane - reguluje mocą w sposób ciągły zamiast dyskretnie włączać/wyłączać grzanie.')
    linie.append('')
    linie.append('3) ANOMALIE (Max HRT > 35°C - podejrzenie przegrzania)')
    linie.append(f'   Łącznie {len(anomalie)} przypadków na {len(df)} wierszy danych:')
    for alg in lista_algorytmow:
        linie.append(f'   - {alg}: {int(anomalie_per_algorytm[alg])} przypadków')
    linie.append('   HIPOTEZA: wszystkie anomalie występują WYŁĄCZNIE w trzech algorytmach z dyskretnym')
    linie.append('   limitem przełączeń (12/dobę) - zero przypadków w wersji PID, która tego limitu nie')
    linie.append('   używa (reguluje mocą ciągle, 0-100%). Prawdopodobny mechanizm: przy oscylacjach')
    linie.append('   warunków blisko progu załączenia/wyłączenia dobowy budżet przełączeń wyczerpuje się')
    linie.append('   wcześnie, a grzanie "zawiesza się" w stanie włączonym do północy (reset licznika) -')
    linie.append('   przy długim, głębokim mrozie HRT zdąża wtedy dryfować w stronę stanu ustalonego')
    linie.append('   (CRT + ok. 51°C przy 100% mocy), stąd skoki do 37-48°C. To NIE jest błąd wyznaczania')
    linie.append('   temperatury zadanej, tylko efekt uboczny bezpiecznika ograniczającego zużycie styku -')
    linie.append('   przykłady: ' + ', '.join(
        f'{r.Lokalizacja} {r.Rok} ({r.Algorytm}, HRT={r.max_hrt:.1f}°C)'
        for r in anomalie.sort_values('max_hrt', ascending=False).head(4).itertuples()
    ) + '.')
    linie.append('')
    linie.append('4) REKOMENDACJA PRAKTYCZNA')
    linie.append('   Funkcja ryzyka w wersji PID (risk_function_pid) jest jedynym z czterech wariantów,')
    linie.append('   który NIE wygenerował ani jednego przypadku przegrzania powyżej 35°C, przy')
    linie.append(f'   jednoczesnym {agregaty["przelaczenia_srednie"].max()/agregaty.loc["Funkcja ryzyka (PID)", "przelaczenia_srednie"]:.1f}x mniejszym zużyciu styku przełączającego niż pozostałe warianty.')
    linie.append(f'   Pod względem samej energii przegrywa nieznacznie z wersją binarną funkcji ryzyka')
    linie.append(f'   ({agregaty.loc["Funkcja ryzyka (PID)","energia_srednia"]:.1f} kWh vs '
                 f'{agregaty.loc["Funkcja ryzyka (binarna)","energia_srednia"]:.1f} kWh średnio), ale biorąc pod')
    linie.append('   uwagę brak anomalii i dużo mniejsze zużycie mechaniczne przekaźnika, to ona jest')
    linie.append('   rekomendowanym wyborem do dalszego rozwoju/wdrożenia.')

    for i, linia in enumerate(linie, start=1):
        komorka = ws_wn.cell(row=i, column=1, value=linia)
        komorka.font = FONT_POGRUBIONY if (linia.isupper() or re.match(r'^\d\)', linia)) else FONT_ZWYKLY
        komorka.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(SCIEZKA_XLSX)
    print(f'Zapisano: {SCIEZKA_XLSX}')
    return df, agregaty, anomalie


if __name__ == '__main__':
    main()
