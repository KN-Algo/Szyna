# generuj_excel_master.py
#
# Buduje JEDEN skonsolidowany Excel (wyniki/Podsumowanie_MASTER.xlsx) zbierający
# NAGŁÓWKOWE wyniki ze WSZYSTKICH testów w tym projekcie - jedna zakładka na
# test, z kluczowymi liczbami (nie pełnym surowym dumpem - ten zawsze zostaje w
# oryginalnym pliku danego testu, tu tylko odnośnik). Każda zakładka jest
# POMIJANA BEZ BŁĘDU, jeśli dany test jeszcze nie był uruchomiony (brak pliku
# wynikowego) - można więc odpalić ten skrypt osobno w dowolnym momencie, nie
# tylko na końcu uruchom_wszystkie_testy.py.
#
# Uruchomienie: python generuj_excel_master.py

import os
import sys
import glob
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_WYNIKOW_GLOWNY = os.path.join(BASE_DIR, 'wyniki')
SCIEZKA_MASTER = os.path.join(FOLDER_WYNIKOW_GLOWNY, 'Podsumowanie_MASTER.xlsx')

FONT_NAZWA = 'Arial'
FONT_TYTUL = Font(name=FONT_NAZWA, bold=True, size=16)
FONT_NAGLOWEK = Font(name=FONT_NAZWA, bold=True, color='FFFFFF', size=11)
FILL_NAGLOWEK = PatternFill('solid', fgColor='1F4E78')
FONT_ZWYKLY = Font(name=FONT_NAZWA, size=10)
FONT_POGRUBIONY = Font(name=FONT_NAZWA, bold=True, size=10)
FONT_SCIEZKA = Font(name=FONT_NAZWA, italic=True, size=9, color='555555')
OBRAMOWANIE = Border(*(Side(style='thin', color='B7B7B7') for _ in range(4)))


def naglowek(ws, wiersz, kolumny):
    for i, nazwa in enumerate(kolumny, start=1):
        c = ws.cell(row=wiersz, column=i, value=nazwa)
        c.font = FONT_NAGLOWEK
        c.fill = FILL_NAGLOWEK
        c.alignment = Alignment(horizontal='center')


def autoszerokosc(ws, min_szer=10, max_szer=45):
    for kolumna_komorki in ws.columns:
        dlugosc = 0
        litera = None
        for komorka in kolumna_komorki:
            if litera is None:
                litera = getattr(komorka, 'column_letter', None)
            if litera is None:
                continue
            if komorka.value is not None:
                dlugosc = max(dlugosc, len(str(komorka.value)))
        if litera:
            ws.column_dimensions[litera].width = max(min_szer, min(max_szer, dlugosc + 3))


def _tabela(ws, df, wiersz_start, sciezka_zrodlowa=None):
    if sciezka_zrodlowa:
        ws.cell(row=wiersz_start, column=1, value=f'Pełne dane: {sciezka_zrodlowa}').font = FONT_SCIEZKA
        wiersz_start += 2
    naglowek(ws, wiersz_start, list(df.columns))
    for i, wiersz in enumerate(df.itertuples(index=False), start=wiersz_start + 1):
        for j, wartosc in enumerate(wiersz, start=1):
            if isinstance(wartosc, float):
                wartosc = round(wartosc, 3)
            c = ws.cell(row=i, column=j, value=wartosc)
            c.font = FONT_ZWYKLY
            c.border = OBRAMOWANIE
    ostatni = wiersz_start + len(df)
    ws.auto_filter.ref = f'A{wiersz_start}:{ws.cell(row=wiersz_start, column=len(df.columns)).coordinate[:-1]}{ostatni}'
    autoszerokosc(ws)
    return ostatni


def sekcja_glowny_przeglad(wb):
    sciezka = os.path.join(FOLDER_WYNIKOW_GLOWNY, 'przeglad_wielu_lokalizacji', 'PRZEGLAD_ZBIORCZY.csv')
    if not os.path.exists(sciezka):
        return
    df = pd.read_csv(sciezka)
    ws = wb.create_sheet('Glowny_przeglad')
    ws.cell(row=1, column=1, value='Główny przegląd - średnia energia per algorytm').font = FONT_TYTUL
    agg = df.groupby('name').agg(
        energia_srednia_kwh=('energia_kwh', 'mean'),
        przelaczenia_srednie=('przelaczenia', 'mean'),
        max_snieg_globalny_mm=('max_snieg_mm', 'max'),
        iae_srednie=('iae', 'mean') if 'iae' in df.columns else ('energia_kwh', 'mean'),
        liczba_lokalizacji=('lokalizacja', 'nunique'),
    ).reset_index().sort_values('energia_srednia_kwh').rename(columns={'name': 'algorytm'})
    _tabela(ws, agg, 3, sciezka)


def sekcja_wrazliwosc_2lok(wb):
    sciezka = os.path.join(FOLDER_WYNIKOW_GLOWNY, 'wrazliwosc_2lokalizacje', 'WRAZLIWOSC_ZBIORCZY.csv')
    if not os.path.exists(sciezka):
        return
    df = pd.read_csv(sciezka)
    ws = wb.create_sheet('Wrazliwosc_transmitancji')
    ws.cell(row=1, column=1, value='Wrażliwość transmitancji + szum (Abisko/Ojmiakon) - odchylenie energii').font = FONT_TYTUL
    baseline = df[(df['scenariusz'] == 'nominal') & (~df['szum'].astype(bool))].set_index(['lokalizacja', 'name'])['energia_kwh']

    def _odch(w):
        b = baseline.get((w['lokalizacja'], w['name']))
        return None if not b else (w['energia_kwh'] - b) / b * 100.0

    df = df.copy()
    df['odchylenie_energii_pct'] = df.apply(_odch, axis=1)
    agg = df.groupby('name').agg(
        odchylenie_sredni_abs_pct=('odchylenie_energii_pct', lambda s: s.abs().mean()),
        autotest_udany_pct=('autotest_fit_ok', lambda s: s.mean() * 100 if s.notna().any() else None),
        blad_K_sredni_pct=('blad_identyfikacji_K_pct', lambda s: s.abs().mean()),
    ).reset_index().sort_values('odchylenie_sredni_abs_pct').rename(columns={'name': 'algorytm'})
    _tabela(ws, agg, 3, sciezka)


def sekcja_krok_sterowania(wb):
    sciezka = os.path.join(FOLDER_WYNIKOW_GLOWNY, 'wrazliwosc_kroku', 'WRAZLIWOSC_KROKU_ZBIORCZY.csv')
    if not os.path.exists(sciezka):
        return
    df = pd.read_csv(sciezka)
    ws = wb.create_sheet('Krok_sterowania')
    ws.cell(row=1, column=1, value='Wrażliwość na krok sterowania - średnia energia per (algorytm, krok)').font = FONT_TYTUL
    agg = df.groupby(['name', 'krok_s']).agg(
        energia_srednia_kwh=('energia_kwh', 'mean'),
        iae_srednie=('iae', 'mean') if 'iae' in df.columns else ('energia_kwh', 'mean'),
    ).reset_index().sort_values(['name', 'krok_s']).rename(columns={'name': 'algorytm'})
    _tabela(ws, agg, 3, sciezka)


def sekcja_awarie(wb):
    sciezka = os.path.join(FOLDER_WYNIKOW_GLOWNY, 'awarie_czujnikow', 'AWARIE_ZBIORCZY.csv')
    if not os.path.exists(sciezka):
        return
    df = pd.read_csv(sciezka)
    ws = wb.create_sheet('Awarie_czujnikow')
    ws.cell(row=1, column=1, value='Odporność na awarie czujników (bias/szum/rozłączenie) - odchylenie energii').font = FONT_TYTUL
    baseline = df[df['scenariusz_awarii'] == 'brak_awarii'].set_index('algorytm')['energia_kwh']

    def _odch(w):
        b = baseline.get(w['algorytm'])
        return None if not b else (w['energia_kwh'] - b) / b * 100.0

    df = df.copy()
    df['odchylenie_energii_pct'] = df.apply(_odch, axis=1)
    agg = df[df['scenariusz_awarii'] != 'brak_awarii'].groupby('algorytm').agg(
        odchylenie_sredni_abs_pct=('odchylenie_energii_pct', lambda s: s.abs().mean()),
        odchylenie_max_abs_pct=('odchylenie_energii_pct', lambda s: s.abs().max()),
    ).reset_index().sort_values('odchylenie_sredni_abs_pct')
    _tabela(ws, agg, 3, sciezka)


def sekcja_szum(wb):
    sciezka = os.path.join(FOLDER_WYNIKOW_GLOWNY, 'szum_wielu_czujnikow', 'SZUM_ZBIORCZY.csv')
    if not os.path.exists(sciezka):
        return
    df = pd.read_csv(sciezka)
    ws = wb.create_sheet('Szum_wielu_czujnikow')
    ws.cell(row=1, column=1, value='Szum wielu czujników (wiele poziomów/sensorów) - odchylenie energii').font = FONT_TYTUL
    baseline = df[df['scenariusz_szumu'] == 'brak_awarii'].set_index(['lokalizacja', 'algorytm'])['energia_kwh']

    def _odch(w):
        b = baseline.get((w['lokalizacja'], w['algorytm']))
        return None if not b else (w['energia_kwh'] - b) / b * 100.0

    df = df.copy()
    df['odchylenie_energii_pct'] = df.apply(_odch, axis=1)
    agg = df[df['scenariusz_szumu'] != 'brak_awarii'].groupby('algorytm').agg(
        odchylenie_sredni_abs_pct=('odchylenie_energii_pct', lambda s: s.abs().mean()),
        odchylenie_max_abs_pct=('odchylenie_energii_pct', lambda s: s.abs().max()),
        liczba_scenariuszy=('scenariusz_szumu', 'nunique'),
    ).reset_index().sort_values('odchylenie_sredni_abs_pct')
    _tabela(ws, agg, 3, sciezka)


def sekcja_prognoza_opadow(wb):
    sciezka = os.path.join(FOLDER_WYNIKOW_GLOWNY, 'Podsumowanie_prognozy_opadow.xlsx')
    if not os.path.exists(sciezka):
        return
    try:
        zrodlo = load_workbook(sciezka, data_only=True)['Wyniki_lokalizacje']
    except Exception:
        return
    ws = wb.create_sheet('Prognoza_opadow')
    ws.cell(row=1, column=1, value='Skuteczność przewidywanie_opadow.py per lokalizacja').font = FONT_TYTUL
    ws.cell(row=3, column=1, value=f'Pełne dane: {sciezka}').font = FONT_SCIEZKA
    for r in range(1, zrodlo.max_row + 1):
        for c in range(1, min(zrodlo.max_column, 9) + 1):  # tylko pierwsze 9 kolumn (bez horyzontu)
            wartosc = zrodlo.cell(row=r, column=c).value
            cel = ws.cell(row=r + 4, column=c, value=wartosc)
            cel.font = FONT_NAGLOWEK if r == 1 else FONT_ZWYKLY
            if r == 1:
                cel.fill = FILL_NAGLOWEK
    autoszerokosc(ws)


def sekcja_indeks(wb, dostepne_sekcje):
    ws = wb.create_sheet('Indeks', 0)
    ws.cell(row=1, column=1, value='Podsumowanie MASTER - indeks wszystkich testów').font = FONT_TYTUL
    ws.cell(row=2, column=1, value='Zbudowane przez generuj_excel_master.py - jedna zakładka na test, '
                                    'pełne dane w oryginalnym pliku każdego testu (ścieżka podana w zakładce).').font = FONT_SCIEZKA

    wszystkie = [
        ('Glowny_przeglad', 'Główny przegląd (test_wszystkie_rownolegle.py)', 'wyniki/przeglad_wielu_lokalizacji/'),
        ('Wrazliwosc_transmitancji', 'Wrażliwość transmitancji + szum, 2 lok. (test_wrazliwosc_dwie_lokalizacje.py)', 'wyniki/wrazliwosc_2lokalizacje/'),
        ('Krok_sterowania', 'Wrażliwość na krok sterowania (test_wrazliwosc_kroku_sterowania.py)', 'wyniki/wrazliwosc_kroku/'),
        ('Awarie_czujnikow', 'Awarie czujników - bias/szum/rozłączenie (test_awarie_czujnikow.py)', 'wyniki/awarie_czujnikow/'),
        ('Szum_wielu_czujnikow', 'Szum wielu czujników, wiele poziomów (test_szum_wielu_czujnikow.py)', 'wyniki/szum_wielu_czujnikow/'),
        ('Prognoza_opadow', 'Skuteczność prognozy opadów (test_skutecznosc_prognozy_opadow.py)', 'wyniki/Podsumowanie_prognozy_opadow.xlsx'),
    ]
    naglowek(ws, 4, ['Zakładka', 'Test', 'Status', 'Folder / plik źródłowy'])
    for i, (klucz, opis, folder) in enumerate(wszystkie, start=5):
        dostepna = klucz in dostepne_sekcje
        ws.cell(row=i, column=1, value=klucz if dostepna else '-').font = FONT_POGRUBIONY
        ws.cell(row=i, column=2, value=opis).font = FONT_ZWYKLY
        status = ws.cell(row=i, column=3, value='dostępne' if dostepna else 'brak danych (test nieuruchomiony)')
        status.font = FONT_ZWYKLY
        if not dostepna:
            status.font = Font(name=FONT_NAZWA, size=10, color='9C0006')
        ws.cell(row=i, column=4, value=folder).font = FONT_SCIEZKA
    autoszerokosc(ws)

    # Diagnostyka funkcji ryzyka ma WŁASNY osobny Excel (jeden na lokalizację,
    # nie jedna zbiorcza tabela CSV) - tu tylko wskazujemy folder, bez tabeli.
    pliki_diag = sorted(glob.glob(os.path.join(FOLDER_WYNIKOW_GLOWNY, 'diagnostyka_funkcji_ryzyka', 'Diagnostyka_*.xlsx')))
    if pliki_diag:
        wiersz = ws.max_row + 3
        ws.cell(row=wiersz, column=1, value='Diagnostyka funkcji ryzyka - osobne pliki Excel (wykresy):').font = FONT_POGRUBIONY
        for j, plik in enumerate(pliki_diag, start=1):
            ws.cell(row=wiersz + j, column=1, value=os.path.relpath(plik, BASE_DIR)).font = FONT_SCIEZKA


def main():
    os.makedirs(FOLDER_WYNIKOW_GLOWNY, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    sekcje = [
        ('Glowny_przeglad', sekcja_glowny_przeglad),
        ('Wrazliwosc_transmitancji', sekcja_wrazliwosc_2lok),
        ('Krok_sterowania', sekcja_krok_sterowania),
        ('Awarie_czujnikow', sekcja_awarie),
        ('Szum_wielu_czujnikow', sekcja_szum),
        ('Prognoza_opadow', sekcja_prognoza_opadow),
    ]
    dostepne = []
    for klucz, funkcja in sekcje:
        try:
            liczba_przed = len(wb.sheetnames)
            funkcja(wb)
            if len(wb.sheetnames) > liczba_przed:
                dostepne.append(klucz)
        except Exception as e:
            print(f"UWAGA: nie udało się zbudować zakładki '{klucz}': {e}")

    sekcja_indeks(wb, dostepne)

    if not dostepne:
        print("UWAGA: żaden test nie ma jeszcze wyników - Excel będzie zawierał tylko pusty indeks. "
              "Uruchom najpierw którykolwiek z testów (albo cały pakiet: uruchom_wszystkie_testy.py).")

    wb.save(SCIEZKA_MASTER)
    print(f"Zapisano: {SCIEZKA_MASTER}")
    print(f"Zakładki z danymi: {', '.join(dostepne) if dostepne else '(brak)'}")


if __name__ == '__main__':
    main()
