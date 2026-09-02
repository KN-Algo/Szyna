# test_szum_wielu_czujnikow.py
#
# PEŁNY test odporności na szum pomiarowy: w odróżnieniu od test_awarie_czujnikow.py
# (3 TYPY awarii - bias/szum/rozłączenie - na 2 czujnikach, 1 lokalizacja, 1 poziom
# szumu) ten skrypt sprawdza WIELE POZIOMÓW białego szumu, na WIELU czujnikach
# (temperatury: HRT/CRT/AT/punkt rosy, wiatr, opad/śnieg), na 10 LOSOWO wybranych
# lokalizacjach (nie wszystkich 43 - świadomy kompromis czas/pokrycie, na życzenie
# użytkownika), dla WSZYSTKICH algorytmów. Cel: dużo szerzej "przeszukać" przestrzeń
# możliwych warunków zaszumienia niż jeden ustalony poziom.
#
# WAŻNE: awaria dotyczy WYŁĄCZNIE tego, co WIDZI kontroler (fault_injector w
# symulacja_fizyczna.uruchom_kontroler, wołany TUŻ PRZED przekazaniem odczytu
# kontrolerowi) - PRAWDZIWA fizyka (rzeczywista temperatura szyny, model śniegu/
# lodu) liczy się zawsze poprawnie z NIEZAFAŁSZOWANYCH wartości - dokładnie tak,
# jak realna awaria czujnika. IAE/ISE/ITAE (jakość regulacji, patrz
# symulacja_fizyczna.uruchom_kontroler) liczone jak zawsze W GŁÓWNEJ PĘTLI
# symulacji, więc automatycznie odzwierciedlają wpływ szumu na regulację.
#
# POZIOMY SZUMU (białego, Gaussa, addytywnego - poza opadem/śniegiem, gdzie
# dodatkowo obcinamy do >=0, bo ujemna intensywność opadu jest fizycznie
# bez sensu) - 4 poziomy na czujnik, kalibrowane osobno wg rzędu wielkości
# każdego sygnału (jawnie udokumentowane, nie "jeden std dla wszystkiego"):
#   temperatury (HRT/CRT/AT/punkt rosy) [°C]: 0.5 / 1.0 / 2.0 / 4.0
#   wiatr [m/s]:                              0.3 / 0.6 / 1.2 / 2.5
#   opad/śnieg [mm/s, po resamplingu]:        0.0001 / 0.0002 / 0.0004 / 0.0008
#     (rząd wielkości: próg "słaby opad" 0.35mm/15min w przewidywanie_opadow.py
#     to ok. 0.00039 mm/s - poziomy szumu są więc porównywalne z realną
#     intensywnością, nie pomijalnie małe ani absurdalnie duże)
#
# 7 czujników x 4 poziomy = 28 scenariuszy szumu + 1 'brak_awarii' (referencja) = 29.
#
# Sterowanie (zmienne środowiskowe):
#   SZYNA_LICZBA_LOKALIZACJI_SZUM - ile losowych lokalizacji (domyślnie 10)
#   SZYNA_SEED_LOKALIZACJI_SZUM   - seed losowania lokalizacji (domyślnie 20260902 - stały, powtarzalny wybór)
#   SZYNA_MAX_DNI_SZUM            - okno dni, najzimniejszy wycinek (domyślnie 10, jak test_awarie_czujnikow.py)
#   SZYNA_KROK_S                  - krok symulacji [s] (domyślnie 10.0)
#   SZYNA_LICZBA_WATKOW           - jak reszta projektu
#   SZYNA_FOLDER_WYNIKOW_SZUM     - folder wyników (domyślnie wyniki/szum_wielu_czujnikow)
#   SZYNA_WZNOW                   - wznawianie (domyślnie 1)
#
# Uruchomienie: python test_szum_wielu_czujnikow.py

import os
import random
import sys
import time
import traceback
from concurrent.futures import ProcessPoolExecutor, as_completed

import numpy as np
import pandas as pd

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
FOLDER_POGODA = os.path.join(BASE_DIR, "Pogoda_pomiary_15_minut")
FOLDER_WYNIKOW = os.environ.get(
    'SZYNA_FOLDER_WYNIKOW_SZUM', os.path.join(BASE_DIR, "wyniki", "szum_wielu_czujnikow"))
os.makedirs(FOLDER_WYNIKOW, exist_ok=True)

LICZBA_LOKALIZACJI = int(os.environ.get('SZYNA_LICZBA_LOKALIZACJI_SZUM', '10'))
SEED_LOKALIZACJI = int(os.environ.get('SZYNA_SEED_LOKALIZACJI_SZUM', '20260902'))
MAX_DNI = int(os.environ.get('SZYNA_MAX_DNI_SZUM', '10'))
KROK_SYMULACJI_S = float(os.environ.get('SZYNA_KROK_S', '10.0'))
MAX_SWITCHES_PER_DAY = 100
WZNAWIAJ_PRZERWANE = os.environ.get('SZYNA_WZNOW', '1') != '0'

_watkow_env = os.environ.get('SZYNA_LICZBA_WATKOW')
LICZBA_WATKOW_NADPISANIE = int(_watkow_env) if _watkow_env else None

_algorytmy_env = os.environ.get('SZYNA_ALGORYTMY')
ALGORYTMY_FILTR = {a.strip() for a in _algorytmy_env.split(',') if a.strip()} if _algorytmy_env else None

SZUM_SEED_BAZA = 20260902  # Bazowy seed generatorów szumu (osobny na sensor+poziom, patrz zbuduj_scenariusze).

# --- CZUJNIKI I POZIOMY SZUMU (patrz nagłówek pliku po uzasadnienie kalibracji) ---
CZUJNIKI_TEMPERATURA = ['HRT_temp_grzana', 'CRT_temp_niegrzana', 'AT_temp_powietrza', 'PUNKT_ROSY_C']
POZIOMY_TEMPERATURA_C = [0.5, 1.0, 2.0, 4.0]

CZUJNIK_WIATR = 'WIATR_M_S'
POZIOMY_WIATR_M_S = [0.3, 0.6, 1.2, 2.5]

CZUJNIKI_OPAD = ['PRECIP_opad', 'SNOW_snieg']
POZIOMY_OPAD_MM_S = [0.0001, 0.0002, 0.0004, 0.0008]

ETYKIETY_POZIOMOW = ['lekki', 'umiarkowany', 'silny', 'ekstremalny']


def wykryj_liczbe_watkow():
    if LICZBA_WATKOW_NADPISANIE:
        return LICZBA_WATKOW_NADPISANIE
    for zmienna in ('SLURM_CPUS_PER_TASK', 'SLURM_JOB_CPUS_PER_NODE', 'PBS_NP', 'NSLOTS', 'LSB_DJOB_NUMPROC'):
        wartosc = os.environ.get(zmienna)
        if wartosc:
            try:
                liczba = int(str(wartosc).split('(')[0].split(',')[0])
                if liczba > 0:
                    return liczba
            except ValueError:
                continue
    try:
        return len(os.sched_getaffinity(0))
    except AttributeError:
        return os.cpu_count() or 1


def wybierz_losowe_lokalizacje():
    wszystkie = sorted(
        os.path.splitext(n)[0] for n in os.listdir(FOLDER_POGODA) if n.endswith('.csv')
    )
    rng = random.Random(SEED_LOKALIZACJI)
    return sorted(rng.sample(wszystkie, min(LICZBA_LOKALIZACJI, len(wszystkie))))


def _zrob_szum_addytywny(pole, std, seed, nieujemny=False):
    rng = np.random.default_rng(seed)

    def _injector(row, index):
        r = dict(row)
        wartosc = r[pole] + float(rng.normal(0.0, std))
        r[pole] = max(0.0, wartosc) if nieujemny else wartosc
        return r
    return _injector


def zbuduj_scenariusze():
    """Zwraca {nazwa_scenariusza: fault_injector_albo_None}. None = brak awarii (baseline)."""
    scenariusze = {'brak_awarii': None}
    seed = SZUM_SEED_BAZA

    for czujnik in CZUJNIKI_TEMPERATURA:
        for poziom_c, etykieta in zip(POZIOMY_TEMPERATURA_C, ETYKIETY_POZIOMOW):
            scenariusze[f'{czujnik}_{etykieta}'] = _zrob_szum_addytywny(czujnik, poziom_c, seed)
            seed += 1

    for poziom_ms, etykieta in zip(POZIOMY_WIATR_M_S, ETYKIETY_POZIOMOW):
        scenariusze[f'{CZUJNIK_WIATR}_{etykieta}'] = _zrob_szum_addytywny(CZUJNIK_WIATR, poziom_ms, seed, nieujemny=True)
        seed += 1

    for czujnik in CZUJNIKI_OPAD:
        for poziom_mms, etykieta in zip(POZIOMY_OPAD_MM_S, ETYKIETY_POZIOMOW):
            scenariusze[f'{czujnik}_{etykieta}'] = _zrob_szum_addytywny(czujnik, poziom_mms, seed, nieujemny=True)
            seed += 1

    return scenariusze


SCENARIUSZE_NAZWY = list(zbuduj_scenariusze().keys())


def przetworz_zadanie(nazwa_lokalizacji, nazwa_algorytmu, nazwa_scenariusza):
    try:
        sys.path.insert(0, os.path.join(BASE_DIR, 'Algorytmy'))
        import symulacja_fizyczna as fiz
        from rejestr_algorytmow import stworz_kontroler, podlega_bezpiecznikowi

        sciezka_csv = os.path.join(FOLDER_POGODA, f"{nazwa_lokalizacji}.csv")
        zakres_dat = fiz.wybierz_najzimniejsze_okno(sciezka_csv, MAX_DNI)
        dt = KROK_SYMULACJI_S
        df_1s = fiz.wczytaj_pogode_1s(sciezka_csv, zakres_dat=zakres_dat, dt=dt)
        A_wd, B_wd, C_wd, D_wd, A_hd, B_hd, C_hd, D_hd, punkty_opoznienia = fiz.przygotuj_modele_stanowe(dt)
        at_array = df_1s['temperatura_powietrza_C'].to_numpy()
        hrt_weather_all = fiz.wylicz_skladowa_pogodowa(at_array, A_wd, B_wd, C_wd, D_wd, dt)

        fault_injector = zbuduj_scenariusze()[nazwa_scenariusza]

        kontroler_normy, metoda_normy = stworz_kontroler('algorytm_z_normy', max_switches_per_day=MAX_SWITCHES_PER_DAY)
        df_normy, stats_normy, snow_ref, power_ref = fiz.uruchom_kontroler(
            'algorytm_z_normy', kontroler_normy, metoda_normy, df_1s, hrt_weather_all,
            A_hd, B_hd, C_hd, D_hd, punkty_opoznienia, dt=dt, print_progress=False,
        )

        if nazwa_algorytmu == 'algorytm_z_normy':
            stats = stats_normy
        else:
            czy_bezpiecznik = podlega_bezpiecznikowi(nazwa_algorytmu)
            kontroler, metoda = stworz_kontroler(nazwa_algorytmu, max_switches_per_day=MAX_SWITCHES_PER_DAY)
            _, stats, _, _ = fiz.uruchom_kontroler(
                nazwa_algorytmu, kontroler, metoda, df_1s, hrt_weather_all,
                A_hd, B_hd, C_hd, D_hd, punkty_opoznienia, dt=dt,
                snow_reference_mm=snow_ref if czy_bezpiecznik else None,
                power_reference_pct=power_ref if czy_bezpiecznik else None,
                print_progress=False, fault_injector=fault_injector,
            )

        stats = dict(stats)
        stats['lokalizacja'] = nazwa_lokalizacji
        stats['algorytm'] = nazwa_algorytmu
        stats['scenariusz_szumu'] = nazwa_scenariusza
        return nazwa_lokalizacji, nazwa_algorytmu, nazwa_scenariusza, stats, None
    except Exception:
        return nazwa_lokalizacji, nazwa_algorytmu, nazwa_scenariusza, None, traceback.format_exc()


def main():
    liczba_watkow = wykryj_liczbe_watkow()
    lokalizacje = wybierz_losowe_lokalizacje()
    sys.path.insert(0, os.path.join(BASE_DIR, 'Algorytmy'))
    from rejestr_algorytmow import ALGORYTMY
    nazwy_algorytmow = list(ALGORYTMY)
    if ALGORYTMY_FILTR is not None:
        nazwy_algorytmow = [a for a in nazwy_algorytmow if a in ALGORYTMY_FILTR]

    print(f"Test szumu wielu czujników: {len(lokalizacje)} losowych lokalizacji (seed={SEED_LOKALIZACJI}): "
          f"{', '.join(lokalizacje)}")
    print(f"{len(nazwy_algorytmow)} algorytmów x {len(SCENARIUSZE_NAZWY)} scenariuszy szumu "
          f"(7 czujników x 4 poziomy + brak_awarii), okno {MAX_DNI} dni, krok {KROK_SYMULACJI_S:g}s, "
          f"{liczba_watkow} procesów.\n")

    zadania = [
        (lok, alg, scen)
        for lok in lokalizacje
        for alg in nazwy_algorytmow
        for scen in SCENARIUSZE_NAZWY
    ]
    liczba_zadan_ogolem = len(zadania)
    print(f"Łącznie {liczba_zadan_ogolem} zadań.\n")

    wyniki = []
    sciezka_zbiorczy = os.path.join(FOLDER_WYNIKOW, "SZUM_ZBIORCZY.csv")
    if WZNAWIAJ_PRZERWANE and os.path.exists(sciezka_zbiorczy):
        try:
            df_poprzedni = pd.read_csv(sciezka_zbiorczy)
            wyniki = df_poprzedni.to_dict('records')
            gotowe = {(w['lokalizacja'], w['algorytm'], w['scenariusz_szumu']) for w in wyniki}
            zadania = [z for z in zadania if z not in gotowe]
            print(f"WZNOWIENIE: {len(gotowe)} gotowych - liczą się tylko brakujące {len(zadania)}/{liczba_zadan_ogolem}.\n")
        except Exception:
            wyniki = []

    if not zadania:
        print("Wszystkie zadania już policzone.")
        if wyniki:
            zbuduj_excel(pd.DataFrame(wyniki))
        return

    bledy = []
    t0 = time.time()
    with ProcessPoolExecutor(max_workers=liczba_watkow) as executor:
        futures = {executor.submit(przetworz_zadanie, lok, alg, scen): (lok, alg, scen)
                   for lok, alg, scen in zadania}
        zakonczone = 0
        for future in as_completed(futures):
            lok, alg, scen, stats, blad = future.result()
            zakonczone += 1
            elapsed_min = (time.time() - t0) / 60.0
            etykieta = f"{lok}/{alg}/{scen}"
            if blad is not None:
                bledy.append((etykieta, blad))
                print(f"[{zakonczone}/{liczba_zadan_ogolem}] BŁĄD {etykieta}:\n{blad}")
            else:
                wyniki.append(stats)
                if zakonczone % 25 == 0 or zakonczone == liczba_zadan_ogolem:
                    print(f"[{zakonczone}/{liczba_zadan_ogolem}] OK ... (upłynęło {elapsed_min:.1f} min)")
            if zakonczone % 25 == 0 or blad is not None:
                pd.DataFrame(wyniki).to_csv(sciezka_zbiorczy, index=False)

    pd.DataFrame(wyniki).to_csv(sciezka_zbiorczy, index=False)
    print(f"\nZakończono w {(time.time() - t0) / 60.0:.1f} min. Sukcesy: {len(wyniki)}/{liczba_zadan_ogolem}. "
          f"Błędy: {len(bledy)}.")
    if bledy:
        print(f"Liczba błędów: {len(bledy)} (pierwsze 10):")
        for etykieta, _ in bledy[:10]:
            print(f"  - {etykieta}")

    if wyniki:
        zbuduj_excel(pd.DataFrame(wyniki))


def zbuduj_excel(df):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule

    FONT_NAZWA = 'Arial'
    FONT_NAGLOWEK = Font(name=FONT_NAZWA, bold=True, color='FFFFFF', size=11)
    FILL_NAGLOWEK = PatternFill('solid', fgColor='1F4E78')
    FONT_ZWYKLY = Font(name=FONT_NAZWA, size=10)
    OBRAMOWANIE = Border(*(Side(style='thin', color='B7B7B7') for _ in range(4)))

    wb = Workbook()

    baseline = df[df['scenariusz_szumu'] == 'brak_awarii'].set_index(['lokalizacja', 'algorytm'])['energia_kwh']

    def _odchylenie(w):
        base = baseline.get((w['lokalizacja'], w['algorytm']))
        if base is None or base == 0:
            return None
        return (w['energia_kwh'] - base) / base * 100.0

    df = df.copy()
    df['odchylenie_energii_pct'] = df.apply(_odchylenie, axis=1)

    ws = wb.active
    ws.title = 'Srednie_per_scenariusz'
    agg = df.groupby(['algorytm', 'scenariusz_szumu']).agg(
        odchylenie_sredni_pct=('odchylenie_energii_pct', lambda s: s.abs().mean()),
        iae_srednie=('iae', 'mean') if 'iae' in df.columns else ('energia_kwh', 'mean'),
    ).reset_index()
    naglowki = ['Algorytm', 'Scenariusz szumu', 'Śr. |odchylenie energii| (%)', 'Śr. IAE (°C·s)']
    for j, n in enumerate(naglowki, start=1):
        c = ws.cell(row=1, column=j, value=n)
        c.font = FONT_NAGLOWEK
        c.fill = FILL_NAGLOWEK
    for i, w in enumerate(agg.itertuples(index=False), start=2):
        ws.cell(row=i, column=1, value=w.algorytm).font = FONT_ZWYKLY
        ws.cell(row=i, column=2, value=w.scenariusz_szumu).font = FONT_ZWYKLY
        odch = ws.cell(row=i, column=3, value=round(w.odchylenie_sredni_pct, 2) if pd.notna(w.odchylenie_sredni_pct) else None)
        odch.font = FONT_ZWYKLY
        iae_c = ws.cell(row=i, column=4, value=round(w.iae_srednie, 1) if pd.notna(w.iae_srednie) else None)
        iae_c.font = FONT_ZWYKLY
    ostatni = len(agg) + 1
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:D{ostatni}'
    skala = ColorScaleRule(start_type='min', start_color='63BE7B', end_type='max', end_color='F8696B')
    ws.conditional_formatting.add(f'C2:C{ostatni}', skala)
    for kol in ('A', 'B', 'C', 'D'):
        ws.column_dimensions[kol].width = 26

    ws2 = wb.create_sheet('Wyniki_surowe')
    kolumny = [c for c in ['lokalizacja', 'algorytm', 'scenariusz_szumu', 'energia_kwh', 'odchylenie_energii_pct',
                            'przelaczenia', 'max_snieg_mm', 'max_hrt', 'min_hrt', 'iae', 'ise', 'itae']
               if c in df.columns]
    for j, n in enumerate(kolumny, start=1):
        c = ws2.cell(row=1, column=j, value=n)
        c.font = FONT_NAGLOWEK
        c.fill = FILL_NAGLOWEK
    for i, wiersz in enumerate(df[kolumny].itertuples(index=False), start=2):
        for j, wartosc in enumerate(wiersz, start=1):
            v = round(wartosc, 3) if isinstance(wartosc, float) and pd.notna(wartosc) else \
                (None if isinstance(wartosc, float) else wartosc)
            cell = ws2.cell(row=i, column=j, value=v)
            cell.font = FONT_ZWYKLY
            cell.border = OBRAMOWANIE
    ws2.freeze_panes = 'A2'
    ws2.auto_filter.ref = f'A1:{ws2.cell(row=1, column=len(kolumny)).coordinate[:-1]}{len(df) + 1}'

    sciezka_xlsx = os.path.join(FOLDER_WYNIKOW, "Podsumowanie_szumu.xlsx")
    wb.save(sciezka_xlsx)
    print(f"Zapisano: {sciezka_xlsx}")


if __name__ == '__main__':
    main()
