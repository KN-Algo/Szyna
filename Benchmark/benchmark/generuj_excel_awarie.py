# generuj_excel_awarie.py
#
# Buduje Podsumowanie_awarii.xlsx z wyników test_awarie_czujnikow.py -
# ODPORNOŚĆ każdego algorytmu na symulowane awarie czujników (bias/szum/
# rozłączenie na HRT i AT) - osobna zakładka "Odpornosc_na_awarie":
# wiersz = algorytm, kolumny = energia/max_hrt/max_snieg dla każdego
# scenariusza awarii + ODCHYLENIE od scenariusza 'brak_awarii' (referencja) -
# duże odchylenie = algorytm źle radzi sobie z tym typem awarii.

import os
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(BASE_DIR, 'Algorytmy'))
from rejestr_algorytmow import ALGORYTMY  # noqa: E402

FONT_NAZWA = 'Arial'
KOLOR_NAGLOWEK_BG = '1F4E78'
KOLOR_NAGLOWEK_FG = 'FFFFFF'
FONT_NAGLOWEK = Font(name=FONT_NAZWA, bold=True, color=KOLOR_NAGLOWEK_FG, size=11)
FILL_NAGLOWEK = PatternFill('solid', fgColor=KOLOR_NAGLOWEK_BG)
FONT_ZWYKLY = Font(name=FONT_NAZWA, size=10)
FONT_POGRUBIONY = Font(name=FONT_NAZWA, bold=True, size=10)
WYROWNANIE_SRODEK = Alignment(horizontal='center', vertical='center')
OBRAMOWANIE_CIENKIE = Border(*(Side(style='thin', color='B7B7B7') for _ in range(4)))

SCENARIUSZE_KOLEJNOSC = ['brak_awarii', 'HRT_bias', 'HRT_szum', 'HRT_rozlaczenie',
                          'AT_bias', 'AT_szum', 'AT_rozlaczenie']


def ustaw_naglowek(ws, wiersz, kolumny):
    for i, nazwa in enumerate(kolumny, start=1):
        komorka = ws.cell(row=wiersz, column=i, value=nazwa)
        komorka.font = FONT_NAGLOWEK
        komorka.fill = FILL_NAGLOWEK
        komorka.alignment = WYROWNANIE_SRODEK
        komorka.border = OBRAMOWANIE_CIENKIE


def autoszerokosc(ws, min_szer=10, max_szer=35):
    for kolumna_komorki in ws.columns:
        dlugosc = 0
        litera = None
        for komorka in kolumna_komorki:
            if litera is None:
                litera = getattr(komorka, 'column_letter', None)
            if litera is None:
                continue
            if komorka.value is not None:
                dlugosc = max(dlugosc, len(str(komorka.value)))
        if litera:
            ws.column_dimensions[litera].width = max(min_szer, min(max_szer, dlugosc + 3))


def main(sciezka_csv, sciezka_xlsx):
    df = pd.read_csv(sciezka_csv)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Odpornosc_na_awarie'
    ws.cell(row=1, column=1, value=(
        'Awaria dotyczy WYŁĄCZNIE tego, co widzi kontroler (fault_injector w symulacja_fizyczna.uruchom_kontroler) - '
        'prawdziwa fizyka szyny/śniegu liczy się zawsze z niezafałszowanych wartości. Kolumny "Δ energia (%)" pokazują '
        'odchylenie od scenariusza "brak_awarii" dla TEGO SAMEGO algorytmu - im większe (dodatnie lub ujemne), tym '
        'gorzej algorytm radzi sobie z danym typem awarii.'
    ))
    ws.cell(row=1, column=1).font = Font(name=FONT_NAZWA, italic=True, size=9, color='555555')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    WIERSZ_NAGLOWKA = 2
    naglowki = ['Algorytm'] + SCENARIUSZE_KOLEJNOSC
    ustaw_naglowek(ws, WIERSZ_NAGLOWKA, naglowki)

    algorytmy_obecne = [a for a in ALGORYTMY if a in df['algorytm'].unique()]
    scenariusze_obecne = [s for s in SCENARIUSZE_KOLEJNOSC if s in df['scenariusz_awarii'].unique()]

    wiersz = WIERSZ_NAGLOWKA + 1
    for alg in algorytmy_obecne:
        ws.cell(row=wiersz, column=1, value=alg).font = FONT_POGRUBIONY
        podzbior = df[df['algorytm'] == alg].set_index('scenariusz_awarii')
        energia_baza = podzbior['energia_kwh'].get('brak_awarii')
        for j, scen in enumerate(scenariusze_obecne, start=2):
            if scen not in podzbior.index:
                continue
            energia = podzbior.loc[scen, 'energia_kwh']
            if scen == 'brak_awarii' or energia_baza is None or energia_baza == 0:
                wartosc_wyswietlana = round(float(energia), 1)
            else:
                delta_pct = (float(energia) - float(energia_baza)) / float(energia_baza) * 100.0
                wartosc_wyswietlana = round(delta_pct, 1)
            komorka = ws.cell(row=wiersz, column=j, value=wartosc_wyswietlana)
            komorka.font = FONT_ZWYKLY
            komorka.border = OBRAMOWANIE_CIENKIE
            komorka.number_format = '0.0"%"' if scen != 'brak_awarii' else '0.0'
        wiersz += 1

    ostatni_wiersz = wiersz - 1
    for j, scen in enumerate(scenariusze_obecne[1:], start=3):  # pomijamy 'brak_awarii' (kolumna referencyjna, bez skali)
        litera = get_column_letter(j)
        skala = ColorScaleRule(start_type='min', start_color='63BE7B', end_type='max', end_color='F8696B')
        ws.conditional_formatting.add(f'{litera}{WIERSZ_NAGLOWKA + 1}:{litera}{ostatni_wiersz}', skala)

    ws.freeze_panes = f'B{WIERSZ_NAGLOWKA + 1}'
    ws.cell(row=WIERSZ_NAGLOWKA, column=1).alignment = WYROWNANIE_SRODEK
    autoszerokosc(ws)

    wb.save(sciezka_xlsx)
    print(f'Zapisano: {sciezka_xlsx}')


if __name__ == '__main__':
    sciezka_csv_domyslna = os.path.join(BASE_DIR, "wyniki", "awarie_czujnikow", "AWARIE_ZBIORCZY.csv")
    sciezka_xlsx_domyslna = os.path.join(BASE_DIR, "wyniki", "awarie_czujnikow", "Podsumowanie_awarii.xlsx")
    main(sciezka_csv_domyslna, sciezka_xlsx_domyslna)
